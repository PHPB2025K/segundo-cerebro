#!/usr/bin/env python3
"""
Enhance Consolidado Financeiro HTML Report - v2
Adds: Margin per product, Cash flow timeline, Automatic alerts, Ads placeholder
"""

import os
import re
import json
import subprocess
import sys
from datetime import datetime, timedelta
from collections import defaultdict
import statistics

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

# ── Paths ──
WORKSPACE = '/root/.openclaw/workspace'
REPORTS = os.path.join(WORKSPACE, 'reports')
HTML_INPUT = os.path.join(REPORTS, 'consolidado-marketplaces-2026-03.html')
HTML_OUTPUT = os.path.join(REPORTS, 'consolidado-marketplaces-2026-03-v2.html')

SHOPEE_FILE = os.path.join(REPORTS, 'extrato-shopee-todas-2026-03-01-2026-03-20.xlsx')
AMAZON_FILE = os.path.join(REPORTS, 'extrato-amazon-2026-03-01-2026-03-20.xlsx')
ML_FILE = os.path.join(REPORTS, 'extrato-ml-0301-a-0320.xlsx')

SHEETS_ID = '1u74aCdH8VrQ2eK01YUQ8fUMwwb6ZPZXvrTTHoexWtnI'

# ── Utility: format BRL ──
def fmt_brl(value):
    """Format number as R$ brasileiro (ponto para milhar, vírgula para decimal)"""
    if value is None or value == 'N/D':
        return 'N/D'
    try:
        value = float(value)
    except (ValueError, TypeError):
        return 'N/D'
    negative = value < 0
    value = abs(value)
    integer = int(value)
    decimal = round((value - integer) * 100)
    int_str = f'{integer:,}'.replace(',', '.')
    result = f'{int_str},{decimal:02d}'
    if negative:
        result = f'-{result}'
    return result

def fmt_pct(value):
    """Format percentage"""
    if value is None or value == 'N/D':
        return 'N/D'
    try:
        return f'{float(value):.1f}%'
    except (ValueError, TypeError):
        return 'N/D'


# ═══════════════════════════════════════════════════════
# 1. LOAD COST DATA FROM GOOGLE SHEETS
# ═══════════════════════════════════════════════════════
def load_cost_data():
    """Load SKU → cost from Google Sheets for each platform"""
    costs = {}  # {platform: {sku: cost}}
    
    for aba in ['SHOPEE', 'MELI', 'AMAZON']:
        try:
            result = subprocess.run(
                ['gog', 'sheets', 'get', SHEETS_ID, f'{aba}!D7:F200', '-p'],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode != 0:
                print(f'  ⚠ Error reading {aba}: {result.stderr.strip()}')
                continue
            
            platform_costs = {}
            for line in result.stdout.strip().split('\n'):
                if not line.strip():
                    continue
                parts = line.split('\t')
                if len(parts) >= 3:
                    sku = parts[0].strip()
                    cost_str = parts[2].strip()
                    # Parse "R$ 6,61" format
                    cost_str = cost_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                    try:
                        cost = float(cost_str)
                        platform_costs[sku] = cost
                    except ValueError:
                        pass  # Skip rows without valid cost
                elif len(parts) >= 1:
                    # SKU without cost
                    pass
            
            costs[aba] = platform_costs
            print(f'  ✓ {aba}: {len(platform_costs)} SKUs com custo')
        except Exception as e:
            print(f'  ⚠ Error loading {aba}: {e}')
            costs[aba] = {}
    
    return costs


# ═══════════════════════════════════════════════════════
# 2. EXTRACT SKU SALES DATA FROM EXCELS
# ═══════════════════════════════════════════════════════
def extract_amazon_sku_data():
    """Extract SKU-level data from Amazon extract (has SKU in metadata)"""
    sku_data = defaultdict(lambda: {'qty': 0, 'revenue': 0, 'taxes': 0, 'count': 0})
    
    wb = openpyxl.load_workbook(AMAZON_FILE, read_only=True)
    ws = wb['Extrato Amazon']
    
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        tipo = row[6].value  # col G
        if tipo != 'SETTLEMENT':
            continue
        
        valor = row[8].value or 0  # col I - Valor
        taxas = row[11].value or 0  # col L - Taxas
        meta_str = row[18].value  # col S - Metadata
        
        if meta_str:
            try:
                meta = json.loads(meta_str)
                sku = meta.get('sku', '')
                qty = meta.get('qty', 1)
                if sku:
                    sku_data[sku]['qty'] += qty
                    sku_data[sku]['revenue'] += float(valor)
                    sku_data[sku]['taxes'] += abs(float(taxas))
                    sku_data[sku]['count'] += 1
            except (json.JSONDecodeError, TypeError):
                pass
    
    wb.close()
    return dict(sku_data)


def extract_cashflow_data():
    """Extract cash flow data from all platforms"""
    cashflow = {
        'Shopee': defaultdict(lambda: {'recebido': 0, 'pendente': 0, 'cancelado': 0, 'recebido_count': 0, 'pendente_count': 0, 'cancelado_count': 0}),
        'Mercado Livre': defaultdict(lambda: {'recebido': 0, 'pendente': 0, 'cancelado': 0, 'recebido_count': 0, 'pendente_count': 0, 'cancelado_count': 0}),
        'Amazon': defaultdict(lambda: {'recebido': 0, 'pendente': 0, 'cancelado': 0, 'recebido_count': 0, 'pendente_count': 0, 'cancelado_count': 0}),
    }
    
    def parse_date(date_val):
        """Parse date from various formats"""
        if date_val is None:
            return None
        if isinstance(date_val, datetime):
            return date_val.date()
        date_str = str(date_val).strip()
        if not date_str:
            return None
        for fmt in ['%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(date_str.split('.')[0], fmt).date()
            except ValueError:
                continue
        return None
    
    def get_week_label(d):
        """Get week label for a date in March 2026"""
        if d is None:
            return None
        day = d.day if hasattr(d, 'day') else d
        if isinstance(d, str):
            return None
        if d.month != 3 or d.year != 2026:
            # Include Feb dates that appear
            if d.month == 2 and d.year == 2026:
                return '24 Fev – 01 Mar'
            return None
        if day <= 7:
            return '01 – 07 Mar'
        elif day <= 14:
            return '08 – 14 Mar'
        elif day <= 20:
            return '15 – 20 Mar'
        else:
            return '21 – 31 Mar'
    
    # Shopee - 3 contas
    wb = openpyxl.load_workbook(SHOPEE_FILE, read_only=True)
    for sheet_name in ['Budamix Store', 'Budamix Oficial', 'Budamix Shop']:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
            valor_liq = row[12].value  # col M
            data_liq = row[14].value  # col O
            status = row[15].value  # col P
            
            if valor_liq is None:
                valor_liq = 0
            
            d = parse_date(data_liq)
            week = get_week_label(d) if d else None
            
            status_str = str(status).strip().lower() if status else ''
            
            if 'cancelado' in status_str:
                if week:
                    cashflow['Shopee'][week]['cancelado'] += abs(float(valor_liq))
                    cashflow['Shopee'][week]['cancelado_count'] += 1
                else:
                    cashflow['Shopee']['Sem data']['cancelado'] += abs(float(valor_liq))
                    cashflow['Shopee']['Sem data']['cancelado_count'] += 1
            elif 'recebido' in status_str or 'approved' in status_str:
                if week:
                    cashflow['Shopee'][week]['recebido'] += float(valor_liq)
                    cashflow['Shopee'][week]['recebido_count'] += 1
            elif 'pendente' in status_str or 'pending' in status_str:
                if week:
                    cashflow['Shopee'][week]['pendente'] += float(valor_liq)
                    cashflow['Shopee'][week]['pendente_count'] += 1
    wb.close()
    
    # Amazon
    wb = openpyxl.load_workbook(AMAZON_FILE, read_only=True)
    ws = wb['Extrato Amazon']
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        valor_liq = row[12].value or 0  # col M
        data_liq = row[14].value  # col O
        status = row[15].value  # col P
        
        d = parse_date(data_liq)
        week = get_week_label(d) if d else None
        
        status_str = str(status).strip().lower() if status else ''
        
        if 'cancelado' in status_str:
            if week:
                cashflow['Amazon'][week]['cancelado'] += abs(float(valor_liq))
                cashflow['Amazon'][week]['cancelado_count'] += 1
        elif 'recebido' in status_str:
            if week:
                cashflow['Amazon'][week]['recebido'] += float(valor_liq)
                cashflow['Amazon'][week]['recebido_count'] += 1
        elif 'pendente' in status_str:
            if week:
                cashflow['Amazon'][week]['pendente'] += float(valor_liq)
                cashflow['Amazon'][week]['pendente_count'] += 1
    wb.close()
    
    # ML
    wb = openpyxl.load_workbook(ML_FILE, read_only=True)
    ws = wb['Extrato 01-20 Mar']
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        valor_liq = row[12].value or 0  # col M
        data_liq = row[14].value  # col O
        status = row[15].value  # col P
        
        d = parse_date(data_liq)
        week = get_week_label(d) if d else None
        
        status_str = str(status).strip().lower() if status else ''
        
        if 'cancelado' in status_str:
            if week:
                cashflow['Mercado Livre'][week]['cancelado'] += abs(float(valor_liq))
                cashflow['Mercado Livre'][week]['cancelado_count'] += 1
        elif 'recebido' in status_str:
            if week:
                cashflow['Mercado Livre'][week]['recebido'] += float(valor_liq)
                cashflow['Mercado Livre'][week]['recebido_count'] += 1
        elif 'pendente' in status_str:
            if week:
                cashflow['Mercado Livre'][week]['pendente'] += float(valor_liq)
                cashflow['Mercado Livre'][week]['pendente_count'] += 1
    wb.close()
    
    return cashflow


# ═══════════════════════════════════════════════════════
# 3. PARSE ABC DATA FROM EXISTING HTML
# ═══════════════════════════════════════════════════════
def parse_abc_from_html(html):
    """Extract ABC product data from the existing HTML"""
    products = []
    
    # Parse Class A products
    # Pattern: rank number, product name, SKU code, qty, revenue, %, acum%, platforms
    class_a_pattern = re.compile(
        r'<span class="rank-number[^"]*">(\d+)</span></td>\s*'
        r'<td>([^<]+)</td>\s*'
        r'<td><code>([^<]+)</code></td>\s*'
        r'<td[^>]*>([0-9.,]+)</td>\s*'
        r'<td[^>]*>([0-9.,]+)</td>\s*'
        r'<td[^>]*>([0-9.,]+)%</td>\s*'
        r'<td[^>]*>([0-9.,]+)%</td>\s*'
        r'<td>(.*?)</td>',
        re.DOTALL
    )
    
    for m in class_a_pattern.finditer(html):
        rank = int(m.group(1))
        name = m.group(2).strip()
        sku = m.group(3).strip()
        qty = int(m.group(4).replace('.', '').replace(',', ''))
        revenue_str = m.group(5).replace('.', '').replace(',', '.')
        revenue = float(revenue_str)
        pct = float(m.group(6).replace(',', '.'))
        acum = float(m.group(7).replace(',', '.'))
        
        # Determine platforms from badges
        platforms_html = m.group(8)
        platforms = []
        if 'badge-green' in platforms_html:
            platforms.append('Shopee')
        if 'badge-blue' in platforms_html:
            platforms.append('ML')
        if 'badge-purple' in platforms_html:
            platforms.append('Amazon')
        
        products.append({
            'rank': rank,
            'name': name,
            'sku': sku,
            'qty': qty,
            'revenue': revenue,
            'pct': pct,
            'acum': acum,
            'platforms': platforms,
            'class': 'A'
        })
    
    # Parse Class B - find the actual table section
    class_b_start = html.find('Classe B — 6 SKUs')
    if class_b_start > 0:
        class_b_end = html.find('</table>', class_b_start) + len('</table>')
        class_b_text = html[class_b_start:class_b_end]
        
        # Extract each row: find <tr> blocks in tbody
        row_pattern = re.compile(r'<tr>\s*<td>(\d+)</td><td>([^<]+)</td><td><code>([^<]+)</code></td>\s*'
                                 r'<td[^>]*>(\d+)</td><td[^>]*>([0-9.,]+)</td>\s*'
                                 r'<td[^>]*>([0-9.,]+)%</td>\s*'
                                 r'<td>(.*?)</td>\s*</tr>',
                                 re.DOTALL)
        
        for m in row_pattern.finditer(class_b_text):
            rank = int(m.group(1))
            name = m.group(2).strip()
            sku = m.group(3).strip()
            qty = int(m.group(4).replace('.', '').replace(',', ''))
            revenue_str = m.group(5).replace('.', '').replace(',', '.')
            revenue = float(revenue_str)
            pct = float(m.group(6).replace(',', '.'))
            
            platforms_html = m.group(7)
            platforms = []
            if 'badge-green' in platforms_html:
                platforms.append('Shopee')
            if 'badge-blue' in platforms_html:
                platforms.append('ML')
            if 'badge-purple' in platforms_html:
                platforms.append('Amazon')
            
            products.append({
                'rank': rank,
                'name': name,
                'sku': sku,
                'qty': qty,
                'revenue': revenue,
                'pct': pct,
                'acum': None,
                'platforms': platforms,
                'class': 'B'
            })
    
    return products


# ═══════════════════════════════════════════════════════
# 4. CALCULATE MARGINS
# ═══════════════════════════════════════════════════════
def calculate_margins(products, costs, amazon_sku_data):
    """Calculate margins for each product using cost data"""
    for p in products:
        sku = p['sku']
        platforms = p['platforms']
        
        # Find cost - prioritize platform-specific cost
        cost_unit = None
        if 'Amazon' in platforms and sku in costs.get('AMAZON', {}):
            cost_unit = costs['AMAZON'][sku]
        elif 'Shopee' in platforms and sku in costs.get('SHOPEE', {}):
            cost_unit = costs['SHOPEE'][sku]
        elif 'ML' in platforms and sku in costs.get('MELI', {}):
            cost_unit = costs['MELI'][sku]
        else:
            # Try any platform
            for platform_key in ['SHOPEE', 'MELI', 'AMAZON']:
                if sku in costs.get(platform_key, {}):
                    cost_unit = costs[platform_key][sku]
                    break
        
        p['cost_unit'] = cost_unit
        
        if cost_unit is not None:
            p['cost_total'] = cost_unit * p['qty']
            # Estimate taxes as platform average
            # Use known tax rates: Shopee ~32.4%, ML ~29.9%, Amazon ~25.7%
            if 'Amazon' in platforms and sku in amazon_sku_data:
                p['taxes'] = amazon_sku_data[sku]['taxes']
            elif 'Amazon' in platforms:
                p['taxes'] = p['revenue'] * 0.257
            elif 'ML' in platforms:
                p['taxes'] = p['revenue'] * 0.299
            else:
                p['taxes'] = p['revenue'] * 0.324
            
            p['profit'] = p['revenue'] - p['cost_total'] - p['taxes']
            p['margin'] = (p['profit'] / p['revenue'] * 100) if p['revenue'] > 0 else 0
        else:
            p['cost_total'] = None
            p['taxes'] = p['revenue'] * 0.31  # overall average
            p['profit'] = None
            p['margin'] = None
    
    return products


# ═══════════════════════════════════════════════════════
# 5. GENERATE ALERTS
# ═══════════════════════════════════════════════════════
def generate_alerts(products):
    """Generate automatic alerts based on margin and tax rules"""
    alerts = []
    
    # Calculate average tax rate for anomaly detection
    tax_rates = []
    for p in products:
        if p['revenue'] > 0 and p.get('taxes'):
            tax_rates.append(p['taxes'] / p['revenue'] * 100)
    
    avg_tax = statistics.mean(tax_rates) if tax_rates else 31.0
    std_tax = statistics.stdev(tax_rates) if len(tax_rates) > 1 else 5.0
    
    for p in products:
        margin = p.get('margin')
        revenue = p.get('revenue', 0)
        
        if margin is None:
            continue
        
        # 🚨 Critical: high revenue + low margin
        if revenue > 5000 and margin < 15:
            alerts.append({
                'type': 'critical',
                'icon': '🚨',
                'title': f'CRÍTICO: {p["name"]}',
                'desc': f'Faturamento R$ {fmt_brl(revenue)} com margem de apenas {fmt_pct(margin)}. Rever precificação urgente.',
                'sku': p['sku'],
                'badge_class': 'badge-red',
                'border_color': 'var(--accent-red)'
            })
        elif margin < 15:
            alerts.append({
                'type': 'danger',
                'icon': '🔴',
                'title': f'Margem crítica: {p["name"]}',
                'desc': f'Margem de {fmt_pct(margin)} — abaixo do mínimo aceitável (15%). Custo unitário R$ {fmt_brl(p.get("cost_unit"))}.',
                'sku': p['sku'],
                'badge_class': 'badge-red',
                'border_color': 'var(--accent-red)'
            })
        elif margin < 25:
            alerts.append({
                'type': 'warning',
                'icon': '🟡',
                'title': f'Atenção: {p["name"]}',
                'desc': f'Margem de {fmt_pct(margin)} — na zona de atenção (< 25%). Monitorar custos e taxas.',
                'sku': p['sku'],
                'badge_class': 'badge-orange',
                'border_color': 'var(--accent-yellow)'
            })
        
        # Tax anomaly
        if p['revenue'] > 0 and p.get('taxes'):
            tax_pct = p['taxes'] / p['revenue'] * 100
            if tax_pct > avg_tax + 2 * std_tax:
                alerts.append({
                    'type': 'tax_anomaly',
                    'icon': '⚠️',
                    'title': f'Taxa anormal: {p["name"]}',
                    'desc': f'Taxa de {fmt_pct(tax_pct)} — significativamente acima da média ({fmt_pct(avg_tax)}). Investigar cobranças.',
                    'sku': p['sku'],
                    'badge_class': 'badge-orange',
                    'border_color': 'var(--accent-orange)'
                })
    
    # Sort: critical first, then danger, warning, tax_anomaly
    priority = {'critical': 0, 'danger': 1, 'warning': 2, 'tax_anomaly': 3}
    alerts.sort(key=lambda a: priority.get(a['type'], 4))
    
    return alerts


# ═══════════════════════════════════════════════════════
# 6. HTML GENERATORS
# ═══════════════════════════════════════════════════════

def generate_enhanced_abc_class_a(products_a):
    """Generate enhanced Class A table with margin columns"""
    rows_html = ''
    for p in products_a:
        rank_class = f'rank-{p["rank"]}' if p['rank'] <= 3 else 'rank-rest'
        
        cost_unit_str = f'R$ {fmt_brl(p["cost_unit"])}' if p.get('cost_unit') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        cost_total_str = f'R$ {fmt_brl(p["cost_total"])}' if p.get('cost_total') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        profit_str = f'R$ {fmt_brl(p["profit"])}' if p.get('profit') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        
        # Margin color
        margin = p.get('margin')
        if margin is not None:
            if margin >= 25:
                margin_color = 'var(--accent-green)'
                margin_str = fmt_pct(margin)
            elif margin >= 15:
                margin_color = 'var(--accent-yellow)'
                margin_str = fmt_pct(margin)
            else:
                margin_color = 'var(--accent-red)'
                margin_str = fmt_pct(margin)
        else:
            margin_color = 'var(--text-muted)'
            margin_str = 'N/D'
        
        # Profit color
        profit_color = ''
        if p.get('profit') is not None:
            if p['profit'] >= 0:
                profit_color = 'color: var(--accent-green);'
            else:
                profit_color = 'color: var(--accent-red);'
        
        # Platform badges
        badges = ''
        for plat in p.get('platforms', []):
            if plat == 'Shopee':
                badges += '<span class="badge-inline badge-green">Shopee</span> '
            elif plat == 'ML':
                badges += '<span class="badge-inline badge-blue">ML</span> '
            elif plat == 'Amazon':
                badges += '<span class="badge-inline badge-purple">Amazon</span> '
        
        acum_str = f'{p["acum"]:.1f}%'.replace('.', ',') if p.get('acum') else ''
        pct_str = f'{p["pct"]:.1f}%'.replace('.', ',') if p.get('pct') else ''
        
        rows_html += f'''        <tr>
          <td><span class="rank-number {rank_class}">{p['rank']}</span></td>
          <td>{p['name']}</td>
          <td><code>{p['sku']}</code></td>
          <td style="text-align: right; font-weight: 600; color: var(--accent-green);">{p['qty']:,}</td>
          <td style="text-align: right;">{fmt_brl(p['revenue'])}</td>
          <td style="text-align: right;">{cost_unit_str}</td>
          <td style="text-align: right;">{cost_total_str}</td>
          <td style="text-align: right; {profit_color}">{profit_str}</td>
          <td style="text-align: right;"><span style="color: {margin_color}; font-weight: 700;">{margin_str}</span></td>
          <td style="text-align: right; color: var(--accent-yellow);">{pct_str}</td>
          <td style="text-align: right; color: var(--text-muted);">{acum_str}</td>
          <td>{badges}</td>
        </tr>
'''
    
    return f'''  <div class="card-top-accent rainbow">
    <h3 class="mb-16">Classe A — 7 SKUs = 78% do faturamento</h3>
    <p style="color: var(--text-secondary); margin-bottom: 16px; font-size: 0.85rem;">Produtos que concentram a maior parte da receita. Agora com <strong style="color: var(--accent-cyan);">margem real</strong> calculada a partir dos custos de compra.</p>
    <table>
      <thead>
        <tr>
          <th style="width: 30px;">#</th>
          <th>Produto</th>
          <th>SKU</th>
          <th style="text-align: right;">Qtd</th>
          <th style="text-align: right;">Bruto (R$)</th>
          <th style="text-align: right;">Custo Un.</th>
          <th style="text-align: right;">Custo Total</th>
          <th style="text-align: right;">Lucro</th>
          <th style="text-align: right;">Margem</th>
          <th style="text-align: right;">%</th>
          <th style="text-align: right;">Acum.</th>
          <th>Plataformas</th>
        </tr>
      </thead>
      <tbody>
{rows_html}      </tbody>
    </table>
  </div>'''


def generate_enhanced_abc_class_b(products_b):
    """Generate enhanced Class B table with margin columns"""
    rows_html = ''
    for p in products_b:
        cost_unit_str = f'R$ {fmt_brl(p["cost_unit"])}' if p.get('cost_unit') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        cost_total_str = f'R$ {fmt_brl(p["cost_total"])}' if p.get('cost_total') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        profit_str = f'R$ {fmt_brl(p["profit"])}' if p.get('profit') is not None else '<span style="color: var(--text-muted);">N/D</span>'
        
        margin = p.get('margin')
        if margin is not None:
            if margin >= 25:
                margin_color = 'var(--accent-green)'
            elif margin >= 15:
                margin_color = 'var(--accent-yellow)'
            else:
                margin_color = 'var(--accent-red)'
            margin_str = fmt_pct(margin)
        else:
            margin_color = 'var(--text-muted)'
            margin_str = 'N/D'
        
        profit_color = ''
        if p.get('profit') is not None:
            profit_color = 'color: var(--accent-green);' if p['profit'] >= 0 else 'color: var(--accent-red);'
        
        badges = ''
        for plat in p.get('platforms', []):
            if plat == 'Shopee':
                badges += '<span class="badge-inline badge-green">Shopee</span> '
            elif plat == 'ML':
                badges += '<span class="badge-inline badge-blue">ML</span> '
            elif plat == 'Amazon':
                badges += '<span class="badge-inline badge-purple">Amazon</span> '
        
        pct_str = f'{p["pct"]:.1f}%'.replace('.', ',') if p.get('pct') else ''
        
        rows_html += f'''        <tr>
          <td>{p['rank']}</td>
          <td>{p['name']}</td>
          <td><code>{p['sku']}</code></td>
          <td style="text-align: right;">{p['qty']:,}</td>
          <td style="text-align: right;">{fmt_brl(p['revenue'])}</td>
          <td style="text-align: right;">{cost_unit_str}</td>
          <td style="text-align: right;">{cost_total_str}</td>
          <td style="text-align: right; {profit_color}">{profit_str}</td>
          <td style="text-align: right;"><span style="color: {margin_color}; font-weight: 700;">{margin_str}</span></td>
          <td style="text-align: right; color: var(--accent-yellow);">{pct_str}</td>
          <td>{badges}</td>
        </tr>
'''
    
    return f'''  <div class="card-top-accent orange" style="margin-top: 16px;">
    <h3 class="mb-16">Classe B — 6 SKUs = 17% do faturamento</h3>
    <p style="color: var(--text-secondary); margin-bottom: 16px; font-size: 0.85rem;">Produtos intermediários com análise de margem real.</p>
    <table>
      <thead>
        <tr>
          <th style="width: 30px;">#</th>
          <th>Produto</th>
          <th>SKU</th>
          <th style="text-align: right;">Qtd</th>
          <th style="text-align: right;">Bruto (R$)</th>
          <th style="text-align: right;">Custo Un.</th>
          <th style="text-align: right;">Custo Total</th>
          <th style="text-align: right;">Lucro</th>
          <th style="text-align: right;">Margem</th>
          <th style="text-align: right;">%</th>
          <th>Plataformas</th>
        </tr>
      </thead>
      <tbody>
{rows_html}      </tbody>
    </table>
  </div>'''


def generate_cashflow_section(cashflow):
    """Generate cash flow timeline section HTML"""
    
    week_order = ['24 Fev – 01 Mar', '01 – 07 Mar', '08 – 14 Mar', '15 – 20 Mar']
    
    platform_configs = [
        ('Shopee', 'var(--accent-green)', 'badge-green', 'green'),
        ('Mercado Livre', '#3b82f6', 'badge-blue', 'cyan'),
        ('Amazon', 'var(--accent-purple)', 'badge-purple', 'purple'),
    ]
    
    # Calculate totals
    total_recebido = sum(
        sum(cashflow[p][w]['recebido'] for w in cashflow[p])
        for p in cashflow
    )
    total_pendente = sum(
        sum(cashflow[p][w]['pendente'] for w in cashflow[p])
        for p in cashflow
    )
    total_cancelado = sum(
        sum(cashflow[p][w]['cancelado'] for w in cashflow[p])
        for p in cashflow
    )
    
    html = f'''
<!-- ═══════════════════════════════════════════════════════
     SEÇÃO: FLUXO DE CAIXA
     ═══════════════════════════════════════════════════════ -->
<section>
  <div class="section-header">
    <div class="section-number" style="background: linear-gradient(135deg, #22d3ee, #3b82f6);">💸</div>
    <span class="section-title">Fluxo de Caixa — Timeline de Liquidação</span>
    <div class="section-line"></div>
  </div>

  <!-- KPIs do Fluxo -->
  <div class="metric-grid" style="grid-template-columns: repeat(3, 1fr); margin-bottom: 24px;">
    <div class="metric-card" style="border-left: 3px solid var(--accent-green);">
      <div class="metric-value text-green">R$ {fmt_brl(total_recebido)}</div>
      <div class="metric-label">✅ Total Recebido</div>
    </div>
    <div class="metric-card" style="border-left: 3px solid var(--accent-yellow);">
      <div class="metric-value text-yellow">R$ {fmt_brl(total_pendente)}</div>
      <div class="metric-label">⏳ Total Pendente</div>
    </div>
    <div class="metric-card" style="border-left: 3px solid var(--accent-red);">
      <div class="metric-value text-red">R$ {fmt_brl(total_cancelado)}</div>
      <div class="metric-label">❌ Total Cancelado</div>
    </div>
  </div>
'''
    
    for plat_name, color, badge_class, accent in platform_configs:
        data = cashflow[plat_name]
        
        # Platform total
        plat_total_rec = sum(data[w]['recebido'] for w in data)
        plat_total_pend = sum(data[w]['pendente'] for w in data)
        
        html += f'''
  <div class="card-top-accent {accent}" style="margin-bottom: 16px;">
    <h3 class="mb-16"><span class="badge-inline {badge_class}">{plat_name}</span> — Timeline de Liquidação</h3>
    <div style="display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap;">
      <div style="padding: 6px 14px; background: var(--bg-card-elevated); border-radius: 8px; font-size: 0.85rem;">
        <span style="color: var(--text-muted);">Recebido:</span>
        <span style="color: var(--accent-green); font-weight: 600;"> R$ {fmt_brl(plat_total_rec)}</span>
      </div>
      <div style="padding: 6px 14px; background: var(--bg-card-elevated); border-radius: 8px; font-size: 0.85rem;">
        <span style="color: var(--text-muted);">Pendente:</span>
        <span style="color: var(--accent-yellow); font-weight: 600;"> R$ {fmt_brl(plat_total_pend)}</span>
      </div>
    </div>
'''
        
        # Timeline bars for each week
        max_val = max((data[w]['recebido'] for w in week_order if w in data), default=1)
        if max_val == 0:
            max_val = 1
        
        for week in week_order:
            if week not in data:
                continue
            d = data[week]
            rec = d['recebido']
            pend = d['pendente']
            canc = d['cancelado']
            
            bar_width = max(5, rec / max_val * 100) if max_val > 0 else 5
            
            html += f'''    <div style="margin-bottom: 12px;">
      <div style="display: flex; justify-content: space-between; font-size: 0.8rem; color: var(--text-muted); margin-bottom: 4px;">
        <span>📅 {week}</span>
        <span style="color: var(--accent-green);">✅ R$ {fmt_brl(rec)}</span>
      </div>
      <div class="progress-bar">
        <div class="progress-fill" style="width: {bar_width:.0f}%; background: {color};">R$ {fmt_brl(rec)}</div>
      </div>
'''
            if pend > 0:
                html += f'''      <div style="font-size: 0.75rem; color: var(--accent-yellow); margin-top: 2px; padding-left: 8px;">⏳ Pendente: R$ {fmt_brl(pend)} ({d["pendente_count"]} transações)</div>
'''
            if canc > 0:
                html += f'''      <div style="font-size: 0.75rem; color: var(--accent-red); margin-top: 2px; padding-left: 8px;">❌ Cancelado: R$ {fmt_brl(canc)} ({d["cancelado_count"]} transações)</div>
'''
            html += '    </div>\n'
        
        html += '  </div>\n'
    
    html += '</section>\n'
    return html


def generate_alerts_section(alerts):
    """Generate alerts section HTML"""
    if not alerts:
        return '''
<section>
  <div class="section-header">
    <div class="section-number" style="background: linear-gradient(135deg, #22c55e, #3b82f6);">✅</div>
    <span class="section-title">Alertas Automáticos</span>
    <div class="section-line"></div>
  </div>
  <div class="card" style="border-left: 3px solid var(--accent-green);">
    <h3 style="color: var(--accent-green);">✅ Nenhum alerta crítico</h3>
    <p style="color: var(--text-secondary); margin-bottom: 0;">Todos os produtos estão dentro dos limites aceitáveis de margem e taxas.</p>
  </div>
</section>
'''
    
    # Count by type
    critical_count = sum(1 for a in alerts if a['type'] == 'critical')
    danger_count = sum(1 for a in alerts if a['type'] == 'danger')
    warning_count = sum(1 for a in alerts if a['type'] == 'warning')
    tax_count = sum(1 for a in alerts if a['type'] == 'tax_anomaly')
    
    html = f'''
<!-- ═══════════════════════════════════════════════════════
     SEÇÃO: ALERTAS AUTOMÁTICOS
     ═══════════════════════════════════════════════════════ -->
<section>
  <div class="section-header">
    <div class="section-number" style="background: linear-gradient(135deg, #ef4444, #f97316);">⚠</div>
    <span class="section-title">Alertas Automáticos</span>
    <div class="section-line"></div>
  </div>

  <div class="metric-grid" style="grid-template-columns: repeat(4, 1fr); margin-bottom: 24px;">
    <div class="metric-card" style="border-left: 3px solid var(--accent-red);">
      <div class="metric-value text-red">{critical_count}</div>
      <div class="metric-label">🚨 Críticos</div>
    </div>
    <div class="metric-card" style="border-left: 3px solid #ef4444;">
      <div class="metric-value" style="color: #ef4444;">{danger_count}</div>
      <div class="metric-label">🔴 Margem < 15%</div>
    </div>
    <div class="metric-card" style="border-left: 3px solid var(--accent-yellow);">
      <div class="metric-value text-yellow">{warning_count}</div>
      <div class="metric-label">🟡 Margem < 25%</div>
    </div>
    <div class="metric-card" style="border-left: 3px solid var(--accent-orange);">
      <div class="metric-value text-orange">{tax_count}</div>
      <div class="metric-label">⚠️ Taxas Anormais</div>
    </div>
  </div>
'''
    
    for alert in alerts:
        html += f'''  <div class="callout" style="border-left: 4px solid {alert['border_color']}; margin-bottom: 12px;">
    <div style="display: flex; align-items: flex-start; gap: 12px;">
      <span style="font-size: 1.5rem;">{alert['icon']}</span>
      <div>
        <div style="font-weight: 700; color: var(--text-primary); margin-bottom: 4px;">
          {alert['title']} <code style="font-size: 0.75rem;">{alert['sku']}</code>
        </div>
        <div style="color: var(--text-secondary); font-size: 0.9rem;">{alert['desc']}</div>
      </div>
    </div>
  </div>
'''
    
    html += '</section>\n'
    return html


def generate_ads_placeholder_row():
    """Generate the Ads placeholder row for the tax analysis table"""
    return '''      <tr>
        <td>Investimento em Ads</td>
        <td><span class="fee-highlight" style="background: rgba(139,92,246,0.15); color: var(--accent-purple);">R$ 0,00 *</span></td>
        <td><span class="fee-highlight" style="background: rgba(139,92,246,0.15); color: var(--accent-purple);">R$ 0,00 *</span></td>
        <td><span class="fee-highlight" style="background: rgba(139,92,246,0.15); color: var(--accent-purple);">R$ 0,00 *</span></td>
      </tr>'''


def generate_ads_note():
    """Generate the ads integration note"""
    return '''
  <div class="callout callout-info" style="margin-top: 16px; border-left: 4px solid var(--accent-purple);">
    <p style="margin: 0;"><strong style="color: var(--accent-purple);">📢 * Investimento em Ads</strong><br>
    Dados de Ads serão integrados na próxima versão. A estrutura está preparada para receber valores reais de investimento em publicidade por plataforma (Amazon Ads, ML Ads, Shopee Ads).</p>
  </div>'''


# ═══════════════════════════════════════════════════════
# MAIN: ASSEMBLE ENHANCED HTML
# ═══════════════════════════════════════════════════════
def main():
    print('🚀 Enhance Consolidado Financeiro v2')
    print('=' * 50)
    
    # 1. Load HTML
    print('\n📄 Lendo HTML existente...')
    with open(HTML_INPUT, 'r', encoding='utf-8') as f:
        html = f.read()
    print(f'  ✓ {len(html)} caracteres lidos')
    
    # 2. Load cost data
    print('\n💰 Carregando custos do Google Sheets...')
    costs = load_cost_data()
    
    # 3. Extract Amazon SKU data
    print('\n📦 Extraindo dados de SKU da Amazon...')
    amazon_sku_data = extract_amazon_sku_data()
    print(f'  ✓ {len(amazon_sku_data)} SKUs encontrados na Amazon')
    
    # 4. Extract cash flow data
    print('\n💸 Extraindo fluxo de caixa...')
    cashflow = extract_cashflow_data()
    for plat in cashflow:
        total = sum(cashflow[plat][w]['recebido'] for w in cashflow[plat])
        print(f'  ✓ {plat}: R$ {fmt_brl(total)} recebido')
    
    # 5. Parse ABC from HTML
    print('\n📊 Parseando Curva ABC do HTML...')
    products = parse_abc_from_html(html)
    print(f'  ✓ {len(products)} produtos encontrados (A + B)')
    
    # 6. Calculate margins
    print('\n📈 Calculando margens...')
    products = calculate_margins(products, costs, amazon_sku_data)
    for p in products:
        margin_str = fmt_pct(p['margin']) if p['margin'] is not None else 'N/D'
        print(f'  {p["sku"]}: margem {margin_str}')
    
    # 7. Generate alerts
    print('\n⚠️ Gerando alertas...')
    alerts = generate_alerts(products)
    print(f'  ✓ {len(alerts)} alertas gerados')
    for a in alerts:
        print(f'    {a["icon"]} {a["title"]}')
    
    # 8. Build enhanced HTML
    print('\n🔧 Montando HTML aprimorado...')
    
    # Split products by class
    class_a = [p for p in products if p['class'] == 'A']
    class_b = [p for p in products if p['class'] == 'B']
    
    # A) Replace Class A table
    class_a_start = html.find('<div class="card-top-accent rainbow">')
    class_a_end_marker = 'Classe B'
    class_a_end = html.find(class_a_end_marker, class_a_start)
    # Go back to find the closing </div> of class A card
    class_a_end = html.rfind('</div>', class_a_start, class_a_end)
    # Find the end of that div properly
    class_a_section_end = html.find('</div>', class_a_end) + len('</div>')
    
    new_class_a = generate_enhanced_abc_class_a(class_a)
    
    # B) Replace Class B table
    class_b_start = html.find('<div class="card-top-accent orange"', class_a_section_end)
    class_b_end = html.find('</div>', html.find('</table>', class_b_start)) + len('</div>')
    
    new_class_b = generate_enhanced_abc_class_b(class_b)
    
    # Rebuild HTML with replacements
    # First, replace Class A
    html = html[:class_a_start] + new_class_a + html[class_a_section_end:]
    
    # Re-find Class B position after Class A replacement
    class_b_start = html.find('<div class="card-top-accent orange"')
    if class_b_start > 0:
        class_b_end = html.find('</div>', html.find('</table>', class_b_start)) + len('</div>')
        html = html[:class_b_start] + new_class_b + html[class_b_end:]
    
    # C) Insert Alerts section after Curva ABC (before Insights)
    insights_marker = '<!-- ═══════════════════════════════════════════════════════\n     SEÇÃO 7: INSIGHTS & RECOMENDAÇÕES'
    insights_pos = html.find(insights_marker)
    if insights_pos < 0:
        # Try alternative
        insights_pos = html.find('SEÇÃO 7: INSIGHTS')
        if insights_pos > 0:
            insights_pos = html.rfind('<!--', 0, insights_pos)
    
    if insights_pos > 0:
        alerts_html = generate_alerts_section(alerts)
        html = html[:insights_pos] + alerts_html + '\n' + html[insights_pos:]
        print('  ✓ Seção de Alertas inserida')
    else:
        print('  ⚠ Não encontrou marcador de Insights, tentando alternativa...')
        # Insert before footer
        footer_pos = html.find('<!-- Footer -->')
        if footer_pos > 0:
            alerts_html = generate_alerts_section(alerts)
            html = html[:footer_pos] + alerts_html + '\n' + html[footer_pos:]
            print('  ✓ Seção de Alertas inserida (antes do footer)')
    
    # D) Insert Cash Flow section after Amazon Detalhamento (section 5)
    # Find end of Amazon section
    amazon_section_end = html.find('SEÇÃO: CURVA ABC')
    if amazon_section_end > 0:
        amazon_section_end = html.rfind('<!--', 0, amazon_section_end)
        cashflow_html = generate_cashflow_section(cashflow)
        html = html[:amazon_section_end] + cashflow_html + '\n' + html[amazon_section_end:]
        print('  ✓ Seção de Fluxo de Caixa inserida')
    else:
        print('  ⚠ Não encontrou fim da seção Amazon')
    
    # E) Add Ads placeholder row in tax analysis table
    # Find the tax table total row
    tax_total_marker = '<td><strong>Total % do Bruto</strong></td>'
    tax_total_pos = html.find(tax_total_marker)
    if tax_total_pos > 0:
        # Find the <tr> start for the total row
        tr_start = html.rfind('<tr', 0, tax_total_pos)
        ads_row = generate_ads_placeholder_row()
        html = html[:tr_start] + ads_row + '\n      ' + html[tr_start:]
        print('  ✓ Linha de Ads inserida na tabela de taxas')
    
    # Add ads note after the tax table
    tax_table_end = html.find('</table>', html.find('Análise de Taxas'))
    if tax_table_end > 0:
        tax_table_end += len('</table>')
        # Find end of the section (next </section>)
        next_section = html.find('</section>', tax_table_end)
        if next_section > 0:
            ads_note = generate_ads_note()
            html = html[:next_section] + ads_note + '\n' + html[next_section:]
            print('  ✓ Nota de Ads inserida')
    
    # F) Update section numbers if needed (Fluxo de Caixa = 6, Curva ABC = 7, Alertas = 8, Insights = 9)
    # The section numbers in the HTML use visual numbers, not section IDs
    
    # G) Update footer
    html = html.replace(
        'Gerado automaticamente por Kobe',
        'Gerado automaticamente por Kobe • v2 — com margem por produto, fluxo de caixa e alertas'
    )
    
    # 9. Save
    print(f'\n💾 Salvando em {HTML_OUTPUT}...')
    with open(HTML_OUTPUT, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'  ✓ {len(html)} caracteres salvos')
    print(f'\n✅ Relatório v2 gerado com sucesso!')
    print(f'   Arquivo: {HTML_OUTPUT}')


if __name__ == '__main__':
    main()
