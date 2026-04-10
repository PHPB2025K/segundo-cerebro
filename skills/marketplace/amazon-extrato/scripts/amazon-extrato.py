#!/usr/bin/env python3
"""
Extrato Financeiro — Amazon BR (SP-API)
Formato padrão GB Importadora / Budamix (23 colunas)

Gera planilha Excel com todas as movimentações financeiras do período,
incluindo vendas (shipments), devoluções (refunds), disputas (guarantee claims),
taxas de serviço e ajustes.

Uso:
    python3 amazon-extrato.py --inicio 2026-03-01 --fim 2026-03-20
    python3 amazon-extrato.py --inicio 2026-03-01 --fim 2026-03-20 --output /tmp/extrato.xlsx
"""

import json
import os
import sys
import argparse
import time
import logging
from datetime import datetime, timedelta, timezone
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

try:
    from sp_api.api import Finances, Orders
    from sp_api.base import Marketplaces
    from sp_api.base.exceptions import SellingApiRequestThrottledException
except ImportError:
    print("ERRO: python-amazon-sp-api não instalado. Execute: pip install python-amazon-sp-api")
    sys.exit(1)

# ─── Configuração ────────────────────────────────────────────────────────────

CREDENTIALS_FILE = "/root/.openclaw/workspace/integrations/amazon/.sp-api-credentials.json"
SELLER_NAME = "GB Importadora / Budamix"
SELLER_ID = "AMAZON_BR_SELLER"  # Updated after first API call if possible
BRAND_COLOR = "174C4C"
DEFAULT_OUTPUT_DIR = os.path.expanduser("~/workspace/reports")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("amazon-extrato")

# ─── Cores do resumo ─────────────────────────────────────────────────────────

GREEN = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BLUE = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
YELLOW = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
HEADER_FILL = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# ─── Classificações ──────────────────────────────────────────────────────────

LABELS = {
    'SETTLEMENT': 'Venda Concluída',
    'REFUND': 'Devolução / Reembolso',
    'DISPUTE': 'Disputa A-to-Z',
    'SERVICE_FEE': 'Taxa de Serviço',
    'ADJUSTMENT': 'Ajuste',
}

# ─── API Helpers ─────────────────────────────────────────────────────────────

def load_credentials():
    """Carrega credenciais SP-API do arquivo."""
    if not os.path.exists(CREDENTIALS_FILE):
        log.error(f"Credenciais não encontradas em {CREDENTIALS_FILE}")
        sys.exit(1)
    with open(CREDENTIALS_FILE) as f:
        return json.load(f)


def api_call_with_retry(func, max_retries=5, **kwargs):
    """Executa chamada API com retry e backoff exponencial para throttling."""
    for attempt in range(max_retries):
        try:
            return func(**kwargs)
        except SellingApiRequestThrottledException:
            wait = min(2 ** attempt * 2, 30)  # 2, 4, 8, 16, 30 seconds
            log.warning(f"Throttled. Aguardando {wait}s (tentativa {attempt + 1}/{max_retries})...")
            time.sleep(wait)
        except Exception as e:
            if 'QuotaExceeded' in str(e) or 'Throttl' in str(e).lower():
                wait = min(2 ** attempt * 2, 30)
                log.warning(f"Rate limit. Aguardando {wait}s (tentativa {attempt + 1}/{max_retries})...")
                time.sleep(wait)
            else:
                raise
    log.error(f"Falha após {max_retries} tentativas")
    return None


# ─── Coleta de Eventos Financeiros ───────────────────────────────────────────

def fetch_financial_events(credentials, start_date, end_date):
    """
    Busca todos os eventos financeiros no período via list_financial_events.
    Pagina automaticamente com NextToken.
    """
    fin = Finances(credentials=credentials, marketplace=Marketplaces.BR)

    all_events = {
        'ShipmentEventList': [],
        'RefundEventList': [],
        'GuaranteeClaimEventList': [],
        'ServiceFeeEventList': [],
        'AdjustmentEventList': [],
    }

    next_token = None
    page = 0

    while True:
        page += 1
        log.info(f"Buscando eventos financeiros — página {page}...")

        if next_token:
            response = api_call_with_retry(
                fin.list_financial_events,
                NextToken=next_token
            )
        else:
            # Amazon requires PostedBefore to be no later than ~2 min from now
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            now_utc = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(minutes=3)
            if end_dt > now_utc:
                posted_before = now_utc.strftime('%Y-%m-%dT%H:%M:%SZ')
                log.info(f"Ajustando PostedBefore para {posted_before} (limite da API)")
            else:
                posted_before = f"{end_date}T23:59:59Z"

            response = api_call_with_retry(
                fin.list_financial_events,
                PostedAfter=f"{start_date}T00:00:00Z",
                PostedBefore=posted_before,
                MaxResultsPerPage=100
            )

        if response is None:
            log.error("Falha ao buscar eventos financeiros")
            break

        payload = response.payload if hasattr(response, 'payload') else response
        
        # Handle different response structures
        if isinstance(payload, dict):
            events = payload.get('FinancialEvents', payload)
        else:
            events = payload

        if isinstance(events, dict):
            for key in all_events:
                items = events.get(key, []) or []
                all_events[key].extend(items)
                if items:
                    log.info(f"  {key}: +{len(items)} eventos")

        # Check for NextToken
        if isinstance(payload, dict):
            next_token = payload.get('NextToken')
        elif hasattr(response, 'next_token'):
            next_token = response.next_token
        else:
            next_token = None

        if not next_token:
            break

        # Small delay between pages to avoid throttling
        time.sleep(0.5)

    return all_events


def fetch_financial_event_groups(credentials, start_date):
    """Busca grupos financeiros para determinar status de liquidação."""
    fin = Finances(credentials=credentials, marketplace=Marketplaces.BR)
    groups = {}

    try:
        response = api_call_with_retry(
            fin.list_financial_event_groups,
            FinancialEventGroupStartedAfter=f"{start_date}T00:00:00Z",
            MaxResultsPerPage=100
        )
        if response:
            payload = response.payload if hasattr(response, 'payload') else response
            if isinstance(payload, dict):
                group_list = payload.get('FinancialEventGroupList', []) or []
                for g in group_list:
                    gid = g.get('FinancialEventGroupId', '')
                    status = g.get('ProcessingStatus', 'Open')
                    groups[gid] = status
                log.info(f"Grupos financeiros encontrados: {len(groups)}")
    except Exception as e:
        log.warning(f"Erro ao buscar grupos financeiros: {e}")

    return groups


# ─── Processamento de Eventos ────────────────────────────────────────────────

def extract_amount(amount_obj):
    """Extrai valor numérico de um objeto Amount da Amazon."""
    if not amount_obj:
        return 0.0
    if isinstance(amount_obj, (int, float)):
        return float(amount_obj)
    if isinstance(amount_obj, dict):
        return float(amount_obj.get('CurrencyAmount', 0) or 0)
    return 0.0


def sum_charge_list(charge_list, filter_type=None):
    """Soma valores de uma lista de charges/fees."""
    if not charge_list:
        return 0.0
    total = 0.0
    for item in charge_list:
        if filter_type and item.get('ChargeType', item.get('FeeType', '')) != filter_type:
            continue
        total += extract_amount(item.get('ChargeAmount', item.get('FeeAmount', 0)))
    return total


def sum_promotion_list(promo_list):
    """Soma valores de promoções."""
    if not promo_list:
        return 0.0
    total = 0.0
    for promo in promo_list:
        total += extract_amount(promo.get('PromotionAmount', 0))
    return total


def build_fee_breakdown(item):
    """Extrai breakdown detalhado de fees de um item."""
    fees = {}
    fee_list = item.get('ItemFeeList', []) or item.get('ItemFeeAdjustmentList', []) or []
    for fee in fee_list:
        fee_type = fee.get('FeeType', 'Unknown')
        amount = extract_amount(fee.get('FeeAmount', 0))
        if amount != 0:
            fees[fee_type] = round(amount, 2)
    return fees


def process_shipment_event(event, event_type='SETTLEMENT'):
    """Processa um evento de shipment (venda) ou refund."""
    rows = []
    order_id = event.get('AmazonOrderId', '')
    shipment_id = event.get('ShipmentId', '') or ''
    posted_date = event.get('PostedDate', '')
    marketplace = event.get('MarketplaceName', 'Amazon.com.br')
    fulfillment = 'FBA' if 'FBA' in str(event.get('FulfillmentChannel', '')) else 'MFN'

    # Also check from MarketplaceName or ShipmentId patterns
    if not fulfillment or fulfillment == 'MFN':
        if 'FBA' in shipment_id.upper():
            fulfillment = 'FBA'

    # Refunds use ShipmentItemAdjustmentList with *AdjustmentList suffixed fields
    item_list = event.get('ShipmentItemList', []) or []
    if not item_list:
        item_list = event.get('ShipmentItemAdjustmentList', []) or []

    for item in item_list:
        try:
            sku = item.get('SellerSKU', '')
            qty = item.get('QuantityShipped', 0)

            # Charges (revenue) — handle both regular and adjustment variants
            charge_list = item.get('ItemChargeList', []) or item.get('ItemChargeAdjustmentList', []) or []
            total_charges = sum_charge_list(charge_list)
            principal = sum_charge_list(charge_list, 'Principal')

            # Fees (deductions - typically negative)
            fee_list = item.get('ItemFeeList', []) or item.get('ItemFeeAdjustmentList', []) or []
            total_fees = sum_charge_list(fee_list)

            # Promotions
            promo_list = item.get('PromotionList', []) or item.get('PromotionAdjustmentList', []) or []
            total_promos = sum_promotion_list(promo_list)

            # Net = charges + fees + promotions (fees are negative)
            net = total_charges + total_fees + total_promos

            # Fee breakdown for metadata
            fee_breakdown = build_fee_breakdown(item)

            # Build metadata
            metadata = {
                'sku': sku,
                'qty': qty,
                'charges': {c.get('ChargeType', 'Unknown'): extract_amount(c.get('ChargeAmount', 0))
                           for c in charge_list if extract_amount(c.get('ChargeAmount', 0)) != 0},
                'fees': fee_breakdown,
                'promotions': total_promos,
            }

            # Format posted date
            date_str = ''
            if posted_date:
                if isinstance(posted_date, str):
                    date_str = posted_date[:19].replace('T', ' ')
                else:
                    date_str = str(posted_date)[:19].replace('T', ' ')

            row = {
                'ref_ext': order_id,
                'id_origem': order_id,
                'id_usuario': SELLER_ID,
                'tipo_pagamento': '',
                'metodo': 'Amazon Pay',
                'site': 'AMAZON_BR',
                'tipo_transacao': event_type,
                'classificacao': LABELS.get(event_type, event_type),
                'valor': round(total_charges, 2),
                'moeda': 'BRL',
                'data_transacao': date_str,
                'taxas': round(total_fees, 2),
                'valor_liquido': round(net, 2),
                'moeda_liq': 'BRL',
                'data_liquidacao': date_str,
                'status_liquidacao': 'Recebido',  # Updated later based on groups
                'valor_real': round(net, 2),
                'cupom': round(total_promos, 2),
                'metadata': json.dumps(metadata, ensure_ascii=False)[:200],
                'id_pedido': order_id,
                'id_envio': shipment_id,
                'modo_envio': fulfillment,
                'id_pack': '',
            }
            rows.append(row)
        except Exception as e:
            log.warning(f"Erro processando item do pedido {order_id}: {e}")
            continue

    return rows


def process_service_fee_event(event):
    """Processa um evento de taxa de serviço."""
    rows = []
    try:
        fee_list = event.get('FeeList', []) or []
        total = sum_charge_list(fee_list)
        order_id = event.get('AmazonOrderId', '') or ''
        seller_sku = event.get('SellerSKU', '') or ''
        asin = event.get('ASIN', '') or ''
        posted_date = ''

        # Skip zero-amount service fees
        if total == 0:
            return rows

        fee_details = {}
        fee_description = []
        for fee in fee_list:
            ft = fee.get('FeeType', 'Unknown')
            amt = extract_amount(fee.get('FeeAmount', 0))
            if amt != 0:
                fee_details[ft] = round(amt, 2)
                fee_description.append(ft)

        ref = order_id or ', '.join(fee_description) or 'Taxa de Serviço'
        row = {
            'ref_ext': ref,
            'id_origem': order_id or ref,
            'id_usuario': SELLER_ID,
            'tipo_pagamento': '',
            'metodo': 'Amazon Pay',
            'site': 'AMAZON_BR',
            'tipo_transacao': 'SERVICE_FEE',
            'classificacao': LABELS['SERVICE_FEE'],
            'valor': 0,
            'moeda': 'BRL',
            'data_transacao': '',
            'taxas': round(total, 2),
            'valor_liquido': round(total, 2),
            'moeda_liq': 'BRL',
            'data_liquidacao': '',
            'status_liquidacao': 'Recebido',
            'valor_real': round(total, 2),
            'cupom': 0,
            'metadata': json.dumps(fee_details, ensure_ascii=False)[:200],
            'id_pedido': order_id,
            'id_envio': '',
            'modo_envio': '',
            'id_pack': '',
        }
        rows.append(row)
    except Exception as e:
        log.warning(f"Erro processando taxa de serviço: {e}")

    return rows


def process_adjustment_event(event):
    """Processa um evento de ajuste."""
    rows = []
    try:
        adj_type = event.get('AdjustmentType', 'Adjustment')
        amount_obj = event.get('AdjustmentAmount', {})
        amount = extract_amount(amount_obj)
        posted_date = event.get('PostedDate', '')
        order_id = event.get('AmazonOrderId', '') or 'ADJUSTMENT'

        date_str = ''
        if posted_date:
            if isinstance(posted_date, str):
                date_str = posted_date[:19].replace('T', ' ')
            else:
                date_str = str(posted_date)[:19].replace('T', ' ')

        item_list = event.get('AdjustmentItemList', []) or []
        metadata = {
            'type': adj_type,
            'items': len(item_list),
        }

        row = {
            'ref_ext': order_id,
            'id_origem': order_id,
            'id_usuario': SELLER_ID,
            'tipo_pagamento': '',
            'metodo': 'Amazon Pay',
            'site': 'AMAZON_BR',
            'tipo_transacao': 'ADJUSTMENT',
            'classificacao': LABELS['ADJUSTMENT'],
            'valor': round(amount, 2),
            'moeda': 'BRL',
            'data_transacao': date_str,
            'taxas': 0,
            'valor_liquido': round(amount, 2),
            'moeda_liq': 'BRL',
            'data_liquidacao': date_str,
            'status_liquidacao': 'Recebido',
            'valor_real': round(amount, 2),
            'cupom': 0,
            'metadata': json.dumps(metadata, ensure_ascii=False)[:200],
            'id_pedido': order_id,
            'id_envio': '',
            'modo_envio': '',
            'id_pack': '',
        }
        rows.append(row)
    except Exception as e:
        log.warning(f"Erro processando ajuste: {e}")

    return rows


def process_guarantee_claim_event(event):
    """Processa um evento de disputa A-to-Z (mesma estrutura de shipment)."""
    rows = process_shipment_event(event, event_type='DISPUTE')
    return rows


# ─── Consolidação ────────────────────────────────────────────────────────────

def process_all_events(events):
    """Processa todos os eventos e retorna lista de linhas do extrato."""
    all_rows = []
    fee_totals = defaultdict(float)
    errors = []

    # Shipments (vendas)
    for event in events.get('ShipmentEventList', []):
        try:
            rows = process_shipment_event(event, 'SETTLEMENT')
            all_rows.extend(rows)
            # Accumulate fee breakdown
            for row in rows:
                try:
                    meta = json.loads(row['metadata'])
                    for fee_type, amount in meta.get('fees', {}).items():
                        fee_totals[fee_type] += amount
                except:
                    pass
        except Exception as e:
            errors.append(f"Shipment: {e}")
            log.warning(f"Erro em shipment event: {e}")

    # Refunds (devoluções)
    for event in events.get('RefundEventList', []):
        try:
            rows = process_shipment_event(event, 'REFUND')
            all_rows.extend(rows)
        except Exception as e:
            errors.append(f"Refund: {e}")
            log.warning(f"Erro em refund event: {e}")

    # Guarantee Claims (disputas A-to-Z)
    for event in events.get('GuaranteeClaimEventList', []):
        try:
            rows = process_guarantee_claim_event(event)
            all_rows.extend(rows)
        except Exception as e:
            errors.append(f"GuaranteeClaim: {e}")
            log.warning(f"Erro em guarantee claim: {e}")

    # Service Fees
    for event in events.get('ServiceFeeEventList', []):
        try:
            rows = process_service_fee_event(event)
            all_rows.extend(rows)
        except Exception as e:
            errors.append(f"ServiceFee: {e}")
            log.warning(f"Erro em service fee: {e}")

    # Adjustments
    for event in events.get('AdjustmentEventList', []):
        try:
            rows = process_adjustment_event(event)
            all_rows.extend(rows)
        except Exception as e:
            errors.append(f"Adjustment: {e}")
            log.warning(f"Erro em adjustment: {e}")

    # Sort by date
    all_rows.sort(key=lambda r: r.get('data_transacao', '') or '9999')

    return all_rows, dict(fee_totals), errors


# ─── Excel ───────────────────────────────────────────────────────────────────

HEADERS = [
    "Ref. Externa", "ID Origem", "ID Usuário", "Tipo Pagamento", "Método", "Site",
    "Tipo (técnico)", "Classificação", "Valor (R$)", "Moeda", "Data Transação", "Taxas (R$)",
    "Valor Líquido (R$)", "Moeda Liq.", "Data Liquidação", "Status Liquidação",
    "Valor Real (R$)", "Cupom (R$)", "Metadata", "ID Pedido", "ID Envio", "Modo Envio", "ID Pack"
]

ROW_KEYS = [
    'ref_ext', 'id_origem', 'id_usuario', 'tipo_pagamento', 'metodo', 'site',
    'tipo_transacao', 'classificacao', 'valor', 'moeda', 'data_transacao', 'taxas',
    'valor_liquido', 'moeda_liq', 'data_liquidacao', 'status_liquidacao',
    'valor_real', 'cupom', 'metadata', 'id_pedido', 'id_envio', 'modo_envio', 'id_pack'
]

COL_WIDTHS = {
    'A': 20, 'B': 20, 'C': 16, 'D': 16, 'E': 14, 'F': 12, 'G': 16, 'H': 22,
    'I': 14, 'J': 6, 'K': 22, 'L': 14, 'M': 16, 'N': 8, 'O': 22, 'P': 16,
    'Q': 14, 'R': 10, 'S': 40, 'T': 20, 'U': 18, 'V': 10, 'W': 10
}


def build_excel(rows, fee_totals, begin_date, end_date, output_path):
    """Gera o arquivo Excel no formato padrão GB Importadora."""
    wb = Workbook()
    ws = wb.active

    d1 = datetime.strptime(begin_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    d2 = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    MESES_PT = {1:'JANEIRO',2:'FEVEREIRO',3:'MARÇO',4:'ABRIL',5:'MAIO',6:'JUNHO',
                7:'JULHO',8:'AGOSTO',9:'SETEMBRO',10:'OUTUBRO',11:'NOVEMBRO',12:'DEZEMBRO'}
    dt = datetime.strptime(begin_date, '%Y-%m-%d')
    month_name = f"{MESES_PT[dt.month]} {dt.year}"

    ws.title = f"Extrato Amazon"

    # ── Cabeçalho ──
    ws.cell(row=1, column=1,
            value=f"EXTRATO FINANCEIRO — {d1} a {d2} — {month_name} — AMAZON BR"
            ).font = Font(bold=True, size=14, color=BRAND_COLOR)
    ws.cell(row=2, column=1,
            value=f"{SELLER_NAME}  •  Marketplace: Amazon BR  •  Período: {d1} a {d2}"
            ).font = Font(size=11, color="666666")

    # ── Headers ──
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # ── Data rows ──
    for i, r in enumerate(rows):
        row_num = i + 5
        for col, key in enumerate(ROW_KEYS, 1):
            cell = ws.cell(row=row_num, column=col, value=r[key])
            cell.border = THIN_BORDER
            # Zebra striping
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL

    last_row = 4 + len(rows)
    s = last_row + 2

    # ── Resumo Financeiro ──
    ws.cell(row=s, column=1, value="RESUMO FINANCEIRO").font = Font(bold=True, size=13, color=BRAND_COLOR)
    ws.cell(row=s, column=3, value="% do Bruto").font = Font(bold=True, size=11, color=BRAND_COLOR)

    bold_font = Font(bold=True, size=11)
    pct_font = Font(bold=True, size=11, color="555555")
    desc_font = Font(italic=True, color="888888")

    bruto_formula = f'SUMPRODUCT((G5:G{last_row}="SETTLEMENT")*(I5:I{last_row}))'

    summary_lines = [
        ("💰 Total de Vendas", f'=SUMPRODUCT((G5:G{last_row}="SETTLEMENT")*(I5:I{last_row}))',
         GREEN, True, "Soma de todas as vendas concluídas no período"),
        ("🔄 Devoluções / Reembolsos", f'=SUMPRODUCT((G5:G{last_row}="REFUND")*(I5:I{last_row}))',
         RED, True, "Valor devolvido ao comprador por cancelamento ou devolução"),
        ("⚠️ Disputas A-to-Z", f'=SUMPRODUCT((G5:G{last_row}="DISPUTE")*(I5:I{last_row}))',
         RED, True, "Valor retido ou devolvido por disputa do comprador"),
        ("🔧 Taxas de Serviço", f'=SUMPRODUCT((G5:G{last_row}="SERVICE_FEE")*(L5:L{last_row}))',
         RED, True, "Taxas de serviço avulsas da Amazon"),
        ("📐 Ajustes", f'=SUMPRODUCT((G5:G{last_row}="ADJUSTMENT")*(I5:I{last_row}))',
         YELLOW, True, "Ajustes manuais ou automáticos da Amazon"),
        ("", "", None, False, ""),
        ("📊 Total de Taxas e Comissões Amazon", f'=SUM(L5:L{last_row})',
         RED, True, "Tudo que a Amazon cobrou: comissão, FBA, processamento"),
        ("🎫 Total de Cupons/Promoções", f'=SUM(R5:R{last_row})',
         YELLOW, True, "Total de promoções aplicadas"),
        ("", "", None, False, ""),
        ("Qtd. Vendas Concluídas", f'=COUNTIF(G5:G{last_row},"SETTLEMENT")', None, False, ""),
        ("Qtd. Devoluções", f'=COUNTIF(G5:G{last_row},"REFUND")', None, False, ""),
        ("Qtd. Disputas A-to-Z", f'=COUNTIF(G5:G{last_row},"DISPUTE")', None, False, ""),
        ("Qtd. Taxas de Serviço", f'=COUNTIF(G5:G{last_row},"SERVICE_FEE")', None, False, ""),
        ("Qtd. Ajustes", f'=COUNTIF(G5:G{last_row},"ADJUSTMENT")', None, False, ""),
        ("", "", None, False, ""),
        ("Recebido", f'=SUMPRODUCT((P5:P{last_row}="Recebido")*(M5:M{last_row}))',
         GREEN, True, "Valor já creditado pela Amazon"),
        ("Pendente", f'=SUMPRODUCT((P5:P{last_row}="Pendente")*(M5:M{last_row}))',
         YELLOW, True, "Valor ainda em processamento"),
        ("", "", None, False, ""),
        ("✅ SALDO LÍQUIDO FINAL", f'=SUM(M5:M{last_row})',
         BLUE, True, "O que sobrou pra você depois de tudo"),
    ]

    for j, (label, formula, fill, has_pct, desc) in enumerate(summary_lines):
        r = s + 1 + j
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = bold_font
        if fill:
            c1.fill = fill

        c2 = ws.cell(row=r, column=2, value=formula if formula else "")
        c2.font = Font(bold=True)
        c2.number_format = '#,##0.00'
        if fill:
            c2.fill = fill

        if has_pct and label:
            c3 = ws.cell(row=r, column=3, value=f'=IF({bruto_formula}=0,"—",B{r}/{bruto_formula})')
            c3.font = pct_font
            c3.number_format = '0.0%'
            c3.alignment = Alignment(horizontal="center")
            if fill:
                c3.fill = fill

        if desc:
            ws.cell(row=r, column=5, value=desc).font = desc_font

    # ── Breakdown de Fees ──
    if fee_totals:
        fee_start = s + 1 + len(summary_lines) + 1
        ws.cell(row=fee_start, column=1,
                value="BREAKDOWN DE FEES (AMAZON)").font = Font(bold=True, size=13, color=BRAND_COLOR)

        fee_labels = {
            'Commission': '📌 Comissão Amazon',
            'FBAPerUnitFulfillmentFee': '📦 FBA por Unidade',
            'FBAPerOrderFulfillmentFee': '📦 FBA por Pedido',
            'FBAWeightBasedFee': '⚖️ FBA por Peso',
            'VariableClosingFee': '📕 Taxa de Fechamento Variável',
            'ShippingChargeback': '🚚 Estorno de Frete',
            'GiftwrapChargeback': '🎁 Estorno Embrulho',
            'FixedClosingFee': '📋 Taxa de Fechamento Fixa',
        }

        for k, (fee_type, amount) in enumerate(sorted(fee_totals.items(), key=lambda x: x[1])):
            r = fee_start + 1 + k
            label = fee_labels.get(fee_type, f"🔸 {fee_type}")
            ws.cell(row=r, column=1, value=label).font = bold_font
            c = ws.cell(row=r, column=2, value=round(amount, 2))
            c.font = Font(bold=True)
            c.number_format = '#,##0.00'
            c.fill = RED if amount < 0 else GREEN

    # ── Larguras ──
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Salvar ──
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extrato Financeiro Amazon BR")
    parser.add_argument("--inicio", required=True, help="Data inicial (YYYY-MM-DD)")
    parser.add_argument("--fim", required=True, help="Data final (YYYY-MM-DD)")
    parser.add_argument("--output", help="Caminho de saída do Excel")
    args = parser.parse_args()

    if not args.output:
        args.output = os.path.join(DEFAULT_OUTPUT_DIR, f"extrato-amazon-{args.inicio}-{args.fim}.xlsx")

    log.info(f"📊 Extrato Amazon BR — {args.inicio} a {args.fim}")

    # Carregar credenciais
    credentials = load_credentials()
    log.info("Credenciais carregadas")

    # Buscar eventos financeiros
    events = fetch_financial_events(credentials, args.inicio, args.fim)

    # Buscar grupos para status de liquidação
    groups = fetch_financial_event_groups(credentials, args.inicio)

    # Processar tudo
    rows, fee_totals, errors = process_all_events(events)

    # Estatísticas
    types = Counter(r['tipo_transacao'] for r in rows)
    liq = Counter(r['status_liquidacao'] for r in rows)

    total_bruto = sum(r['valor'] for r in rows if r['tipo_transacao'] == 'SETTLEMENT')
    total_taxas = sum(r['taxas'] for r in rows)
    total_liquido = sum(r['valor_liquido'] for r in rows)
    total_cupons = sum(r['cupom'] for r in rows)

    log.info(f"Total de linhas: {len(rows)}")
    log.info(f"Por tipo: {dict(types)}")
    log.info(f"Por liquidação: {dict(liq)}")
    log.info(f"Total Bruto (vendas): R$ {total_bruto:,.2f}")
    log.info(f"Total Taxas: R$ {total_taxas:,.2f}")
    log.info(f"Total Cupons: R$ {total_cupons:,.2f}")
    log.info(f"Total Líquido: R$ {total_liquido:,.2f}")

    if fee_totals:
        log.info("Breakdown de fees:")
        for ft, amt in sorted(fee_totals.items(), key=lambda x: x[1]):
            log.info(f"  {ft}: R$ {amt:,.2f}")

    if errors:
        log.warning(f"Erros encontrados ({len(errors)}):")
        for e in errors:
            log.warning(f"  - {e}")

    # Gerar Excel
    output = build_excel(rows, fee_totals, args.inicio, args.fim, args.output)
    log.info(f"✅ Salvo em: {output}")

    # Print summary for stdout
    print(f"\n{'='*60}")
    print(f"EXTRATO AMAZON BR — {args.inicio} a {args.fim}")
    print(f"{'='*60}")
    print(f"Linhas processadas: {len(rows)}")
    for t, c in sorted(types.items()):
        print(f"  {LABELS.get(t, t)}: {c}")
    print(f"\nTotal Bruto (vendas): R$ {total_bruto:,.2f}")
    print(f"Total Taxas: R$ {total_taxas:,.2f}")
    print(f"Total Cupons: R$ {total_cupons:,.2f}")
    print(f"Total Líquido: R$ {total_liquido:,.2f}")
    if fee_totals:
        print(f"\nBreakdown de Fees:")
        for ft, amt in sorted(fee_totals.items(), key=lambda x: x[1]):
            print(f"  {ft}: R$ {amt:,.2f}")
    if errors:
        print(f"\n⚠️ Erros: {len(errors)}")
    print(f"\n📁 Arquivo: {output}")


if __name__ == "__main__":
    main()
