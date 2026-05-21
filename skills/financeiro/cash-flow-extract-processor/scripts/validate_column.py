"""Valida coluna calculada (BASE_MES + dia) na DFC DIÁRIO REALIZADO.

⚠️ Knowledge File v10.1 §"Erros Comuns #5":
Há coluna "Total" intercalada em cada mês (ex: col 125 = "Total" de maio,
antes do dia 1 na col 126). A fórmula `BASE + dia` continua correta, mas
SEMPRE validar `cell(row=4,col)` = mês e `cell(row=5,col)` = dia antes de
escrever, pra evitar escrever no lugar errado.
"""

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from _constants import BASES_MESES_2026


def validar_coluna(planilha_path: Path, mes: str, dia: int) -> tuple[int, str]:
    """Confirma que coluna calculada bate com o mês/dia esperado.

    Args:
        planilha_path: caminho do DFC DIARIO REALIZADO - 2026.xlsx
        mes: nome do mês em maiúsculas (ex: "MAIO")
        dia: dia do mês (1-31)

    Returns:
        (col, status_msg) — col é o índice da coluna (1-based)
    """
    mes_upper = mes.upper()
    if mes_upper not in BASES_MESES_2026:
        raise ValueError(f"Mês inválido: {mes}. Esperado: {list(BASES_MESES_2026.keys())}")

    base = BASES_MESES_2026[mes_upper]
    col = base + dia

    wb = openpyxl.load_workbook(planilha_path, data_only=False, read_only=True)
    ws = wb["DFC"] if "DFC" in wb.sheetnames else wb.active

    mes_check = ws.cell(row=4, column=col).value
    dia_check = ws.cell(row=5, column=col).value

    if mes_check != mes_upper:
        raise AssertionError(
            f"Coluna {col}: mês na planilha = {mes_check!r}, esperado = {mes_upper!r}"
        )
    if dia_check != dia:
        raise AssertionError(
            f"Coluna {col}: dia na planilha = {dia_check!r}, esperado = {dia}"
        )

    return col, f"✅ coluna {col} = {mes_upper} dia {dia}"


def main():
    if len(sys.argv) != 4:
        print("Uso: python validate_column.py <planilha_diario.xlsx> <MES> <DIA>")
        print("Ex:  python validate_column.py 'DFC DIARIO REALIZADO - 2026.xlsx' MAIO 21")
        sys.exit(1)

    try:
        col, msg = validar_coluna(Path(sys.argv[1]), sys.argv[2], int(sys.argv[3]))
        print(msg)
        sys.exit(0)
    except (AssertionError, ValueError) as e:
        print(f"❌ {e}")
        sys.exit(2)


if __name__ == "__main__":
    main()
