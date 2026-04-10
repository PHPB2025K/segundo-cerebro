#!/usr/bin/env python3
"""
update-ml-return-rates — Atualiza Col O (DEVOLUÇÕES) da aba MELI
com taxas reais de devolução por SKU.

Uso:
  python3 update-return-rates.py <arquivo_xlsx_ou_json> [--dry-run] [--no-telegram]

Aceita:
  - xlsx: taxa_devolucao_MLB_perdas_reais.xlsx (col A=MLB, B=SKU, D=vendas, E=devol)
  - json: {"SKU": {"vendas": N, "devolucoes": N, "taxa": X.X}, ...}
"""

import json
import subprocess
import sys
import os
import time
from collections import defaultdict

# ============================================================
# CONFIG
# ============================================================

SSID = "1u74aCdH8VrQ2eK01YUQ8fUMwwb6ZPZXvrTTHoexWtnI"
SHEET = "MELI"
DATA_START_ROW = 7
DATA_END_ROW = 80  # scan up to row 80 to be safe
DEFAULT_RATE = 4.0  # % default when no data
COL_SKU = "D"
COL_PRICE = "G"
COL_DEVOL = "O"
COL_MARGIN = "T"
COL_PROFIT = "U"

# SKU alias map: xlsx_sku -> sheet_sku
SKU_ALIASES = {
    "IMB501PT": "IMB501P_T",
    "IMB501CT": "IMB501C_T",
    "IMB501VT": "IMB501V_T",
    "KIT2YW1520AZ": "KIT2YW1520",
    "KIT4YW1520AZ": "KIT4YW1520",
    "KIT4YW800SQ_T": "KIT4YW800SQ",
    "KIT2YW800SQ_T": "KIT2YW800SQ",
    "096": "KIT3S096",
    "097": "KIT6S097",
    "099": "KIT3S099",
    "102": "KIT9S098",
    "100": "KIT6S100",
    "101": "KIT9S101",
    "098": "KIT9S098",
    "SPC0111": "SPC011",  # variante dupla -> single
}

# ============================================================
# GOG HELPERS
# ============================================================

def setup_gog_env():
    """Set GOG environment variables from systemd override."""
    try:
        result = subprocess.run(
            ["grep", "GOG_KEYRING_PASSWORD",
             "/root/.config/systemd/user/openclaw-gateway.service.d/override.conf"],
            capture_output=True, text=True
        )
        kp = result.stdout.split("GOG_KEYRING_PASSWORD=")[1].split('"')[0].split("\n")[0].strip()
        os.environ["GOG_KEYRING_PASSWORD"] = kp
        os.environ["GOG_ACCOUNT"] = "gb.ai.agent@gbimportadora.com"
        return True
    except Exception as e:
        print(f"ERRO: Não conseguiu configurar GOG: {e}")
        return False


def gog_get(range_str):
    """Read from Google Sheets."""
    result = subprocess.run(
        ["gog", "sheets", "get", SSID, range_str, "--plain"],
        capture_output=True, text=True, timeout=15
    )
    return result.stdout.strip()


def gog_update(cell, value):
    """Write single cell to Google Sheets."""
    result = subprocess.run(
        ["gog", "sheets", "update", SSID, cell, str(value)],
        capture_output=True, text=True, timeout=15
    )
    ok = "updated" in result.stdout.lower()
    if not ok:
        print(f"  WARN: falha ao atualizar {cell}: {result.stdout.strip()}")
    return ok


# ============================================================
# DATA LOADING
# ============================================================

def load_xlsx(path):
    """Load return rates from xlsx file."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    raw = defaultdict(lambda: {"vendas": 0, "devolucoes": 0})

    for r in range(2, ws.max_row + 1):
        mlb = ws.cell(r, 1).value
        if not mlb or mlb == "TOTAL":
            continue
        sku = ws.cell(r, 2).value
        if not sku:
            sku = f"NO_SKU_{mlb}"
        vendas = ws.cell(r, 4).value or 0
        devol = ws.cell(r, 5).value or 0

        raw[sku]["vendas"] += vendas
        raw[sku]["devolucoes"] += devol

    # Calculate rates and resolve aliases
    result = {}
    for sku, data in raw.items():
        total = data["vendas"] + data["devolucoes"]
        taxa = (data["devolucoes"] / total * 100) if total > 0 else 0

        # Resolve alias
        resolved_sku = SKU_ALIASES.get(sku, sku)

        # If alias points to same target, consolidate
        if resolved_sku in result:
            existing = result[resolved_sku]
            combined_vendas = existing["vendas"] + data["vendas"]
            combined_devol = existing["devolucoes"] + data["devolucoes"]
            combined_total = combined_vendas + combined_devol
            result[resolved_sku] = {
                "vendas": combined_vendas,
                "devolucoes": combined_devol,
                "taxa": round((combined_devol / combined_total * 100) if combined_total > 0 else 0, 1),
            }
        else:
            result[resolved_sku] = {
                "vendas": data["vendas"],
                "devolucoes": data["devolucoes"],
                "taxa": round(taxa, 1),
            }

    return result


def load_json(path):
    """Load return rates from JSON file."""
    with open(path) as f:
        raw = json.load(f)

    result = {}
    for sku, data in raw.items():
        resolved_sku = SKU_ALIASES.get(sku, sku)
        result[resolved_sku] = {
            "vendas": data.get("vendas", 0),
            "devolucoes": data.get("devolucoes", 0),
            "taxa": round(data.get("taxa", 0), 1),
        }
    return result


def load_data(path):
    """Auto-detect format and load."""
    if path.endswith(".json"):
        return load_json(path)
    elif path.endswith(".xlsx"):
        return load_xlsx(path)
    else:
        print(f"ERRO: Formato não suportado: {path}")
        print("Aceita: .xlsx ou .json")
        sys.exit(1)


# ============================================================
# FORMULA BUILDER
# ============================================================

def rate_to_formula(row, taxa):
    """Convert return rate % to locale-safe Google Sheets formula."""
    if taxa == 0:
        return "0"

    # Integer percentage: =G*X%
    if taxa == int(taxa):
        return f"={COL_PRICE}{row}*{int(taxa)}%"

    # Fractional: =G*{permille}/1000 to avoid decimal dot
    permille = round(taxa * 10)
    return f"={COL_PRICE}{row}*{permille}/1000"


# ============================================================
# TELEGRAM REPORTING
# ============================================================

def send_telegram_report(report_text, no_telegram=False):
    """Send report to Telegram Thread 3 (Marketplaces)."""
    if no_telegram:
        print("\n[Telegram skip: --no-telegram]")
        return

    try:
        # Use openclaw telegram send if available
        result = subprocess.run(
            ["openclaw", "telegram", "send", "--thread", "3", "--text", report_text],
            capture_output=True, text=True, timeout=15
        )
        if result.returncode == 0:
            print("\n[Telegram: enviado Thread 3]")
        else:
            # Fallback: try curl to Telegram API
            print(f"\n[Telegram: openclaw falhou, tentando fallback]")
            print(f"  {result.stderr.strip()}")
    except Exception as e:
        print(f"\n[Telegram: erro — {e}]")


# ============================================================
# MAIN
# ============================================================

def main():
    # Parse args
    if len(sys.argv) < 2:
        print("Uso: python3 update-return-rates.py <arquivo.xlsx|.json> [--dry-run] [--no-telegram]")
        sys.exit(1)

    file_path = sys.argv[1]
    dry_run = "--dry-run" in sys.argv
    no_telegram = "--no-telegram" in sys.argv

    if not os.path.exists(file_path):
        print(f"ERRO: Arquivo não encontrado: {file_path}")
        sys.exit(1)

    # Setup
    if not setup_gog_env():
        sys.exit(1)

    # Load data
    print(f"Carregando dados de: {file_path}")
    return_data = load_data(file_path)
    print(f"SKUs no arquivo: {len(return_data)}")

    # Read sheet SKUs
    print(f"\nLendo SKUs da planilha {SHEET}...")
    sheet_skus = []
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        sku = gog_get(f"{SHEET}!{COL_SKU}{row}")
        if not sku or sku.strip() == "":
            continue
        sheet_skus.append({"row": row, "sku": sku.strip()})

    print(f"SKUs na planilha: {len(sheet_skus)}")

    if dry_run:
        print("\n[DRY RUN — nenhuma alteração será feita]")

    # Apply updates
    print(f"\n{'=' * 100}")
    print(f"{'DRY RUN: ' if dry_run else ''}Atualizando Col {COL_DEVOL} (DEVOLUÇÕES)")
    print(f"{'=' * 100}")
    print(f"{'Row':>4} {'SKU':<25} {'Devol Atual':>12} {'Formula Nova':<20} {'Fonte'}")
    print("-" * 90)

    updated_real = 0
    kept_default = 0
    not_found_in_sheet = []

    for item in sheet_skus:
        row = item["row"]
        sku = item["sku"]

        old_devol = gog_get(f"{SHEET}!{COL_DEVOL}{row}")

        if sku in return_data:
            rd = return_data[sku]
            taxa = rd["taxa"]
            vendas = rd["vendas"]
            devol = rd["devolucoes"]
            formula = rate_to_formula(row, taxa)
            source = f"real ({vendas}v/{devol}d={taxa:.1f}%)"

            if not dry_run:
                gog_update(f"{SHEET}!{COL_DEVOL}{row}", formula)

            updated_real += 1
        else:
            formula = f"={COL_PRICE}{row}*{int(DEFAULT_RATE)}%"
            source = f"default {DEFAULT_RATE}% (sem dados)"

            if not dry_run:
                gog_update(f"{SHEET}!{COL_DEVOL}{row}", formula)

            kept_default += 1

        print(f"{row:>4} {sku:<25} {old_devol:>12} {formula:<20} {source}")
        time.sleep(0.3)

    # Check which xlsx SKUs didn't match any sheet SKU
    sheet_sku_set = {s["sku"] for s in sheet_skus}
    unmatched = [sku for sku in return_data.keys()
                 if sku not in sheet_sku_set and not sku.startswith("NO_SKU_")]

    print(f"\n{'=' * 100}")
    print(f"ATUALIZAÇÃO CONCLUÍDA{'  (DRY RUN)' if dry_run else ''}")
    print(f"  Com taxa real: {updated_real}")
    print(f"  Default {DEFAULT_RATE}%: {kept_default}")
    if unmatched:
        print(f"  SKUs no xlsx sem match na planilha ({len(unmatched)}):")
        for sku in sorted(unmatched):
            rd = return_data[sku]
            print(f"    {sku} ({rd['vendas']}v/{rd['devolucoes']}d={rd['taxa']:.1f}%)")

    if dry_run:
        print("\n[DRY RUN completo — nenhuma célula alterada]")
        return

    # Wait for recalculation
    print("\nAguardando recálculo...")
    time.sleep(3)

    # Read back margins
    print(f"\n{'=' * 100}")
    print("MARGENS RECALCULADAS")
    print(f"{'=' * 100}")

    alerts = []
    prejuizo = []
    ok_count = 0
    errors = []

    print(f"\n{'SKU':<25} {'Preço':>8} {'Devol':>8} {'Margem':>8} {'Lucro':>12} {'Taxa':>8} {'Status'}")
    print("-" * 95)

    for item in sheet_skus:
        row = item["row"]
        sku = item["sku"]

        price_str = gog_get(f"{SHEET}!{COL_PRICE}{row}")
        devol_str = gog_get(f"{SHEET}!{COL_DEVOL}{row}")
        margin_str = gog_get(f"{SHEET}!{COL_MARGIN}{row}")
        profit_str = gog_get(f"{SHEET}!{COL_PROFIT}{row}")

        has_error = any(e in margin_str for e in ["#ERROR", "#VALUE", "#REF"])

        try:
            margin = float(margin_str.replace("%", "").replace(",", ".").strip())
        except:
            margin = None

        taxa_str = f"{return_data[sku]['taxa']:.1f}%" if sku in return_data else f"{DEFAULT_RATE}%*"

        if has_error:
            status = "ERROR"
            errors.append({"sku": sku, "row": row, "margin": margin_str})
        elif margin is not None and margin < 0:
            status = "PREJUIZO"
            prejuizo.append({"sku": sku, "margin": margin, "profit": profit_str})
        elif margin is not None and margin < 5:
            status = "ALERTA"
            alerts.append({"sku": sku, "margin": margin, "profit": profit_str})
        else:
            status = "OK"
            ok_count += 1

        print(f"{sku:<25} {price_str:>8} {devol_str:>8} {margin_str:>8} {profit_str:>12} {taxa_str:>8} {status}")

    # Summary
    print(f"\n{'=' * 100}")
    print("RESUMO FINAL")
    print(f"  OK (margem >= 5%): {ok_count}")
    print(f"  ALERTA (0-5%):     {len(alerts)}")
    print(f"  PREJUIZO (<0%):    {len(prejuizo)}")
    if errors:
        print(f"  ERRORS:            {len(errors)}")

    if alerts:
        print(f"\nALERTAS:")
        for a in alerts:
            print(f"  {a['sku']:<25} margem {a['margin']:>5.1f}% | {a['profit']}")

    if prejuizo:
        print(f"\nPREJUIZO:")
        for p in prejuizo:
            print(f"  {p['sku']:<25} margem {p['margin']:>5.1f}% | {p['profit']}")

    # Build Telegram report
    tg = []
    tg.append("MELI Devoluções — Atualização de Taxas Reais")
    tg.append(f"Fonte: {os.path.basename(file_path)}")
    tg.append(f"SKUs atualizados (taxa real): {updated_real}")
    tg.append(f"SKUs mantidos (default {DEFAULT_RATE}%): {kept_default}")
    tg.append(f"OK: {ok_count} | Alerta: {len(alerts)} | Prejuizo: {len(prejuizo)}")

    if prejuizo:
        tg.append(f"\nPREJUIZO ({len(prejuizo)} SKUs):")
        for p in sorted(prejuizo, key=lambda x: x["margin"])[:10]:
            tg.append(f"  {p['sku']} margem {p['margin']:.0f}% {p['profit']}")

    if unmatched:
        tg.append(f"\nSKUs no arquivo sem match ({len(unmatched)}): {', '.join(sorted(unmatched)[:5])}")

    send_telegram_report("\n".join(tg), no_telegram)


if __name__ == "__main__":
    main()
