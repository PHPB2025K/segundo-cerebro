#!/usr/bin/env python3
import json, csv, urllib.request, urllib.error, urllib.parse, os, re, subprocess, sys, time
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from openpyxl import load_workbook

ROOT=Path('/root/segundo-cerebro')
OUTDIR=ROOT/'outputs'/'nfs-maio-2026-matriz-filial-final'
OUTDIR.mkdir(parents=True, exist_ok=True)
RATEIO=ROOT/'outputs'/'rateio-fiscal-maio-potes-cnpjs-simples-2026-06-08.xlsx'
RECON=ROOT/'outputs'/'reconciliacao-matriz-filial-maio-2026'/'reconciliacao-matriz-filial-maio-2026-refinada.xlsx'
MATRIZ_PREV=ROOT/'outputs'/'bling-nfe-matriz-maio-cnpjs-simples-2026-06-09.json'
FILIAL_PREV_DIR=ROOT/'tmp'/'bling_payloads_abril'

ACCOUNTS={
 'matriz': {
   'token_path': Path('/root/.openclaw/workspace/scripts/bling-oauth/tokens-matriz.json'),
   'natureza': 15106994870,
   'label':'Matriz',
 },
 'filial': {
   'token_path': Path('/root/.openclaw/workspace/scripts/bling-oauth/tokens-filial.json'),
   'natureza': 15107403191,
   'label':'Filial',
 }
}

DEST_FILE={'GB Comércio':'GB_Comercio.json','Trades':'Trades.json','Broglio':'Broglio.json'}

def brt_now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
def D(x): return Decimal(str(x))
def money(x): return D(x).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def load_rateio_rows():
    wb=load_workbook(RATEIO,data_only=True); ws=wb['Detalhe por SKU']
    rows=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        dest,cnpj,grupo,sku,desc,un,custo,preco,valor,icms=r
        sku2='YW320SQ' if sku=='YW320' else sku
        rows.append(dict(dest=dest,cnpj=cnpj,grupo=grupo,sku=sku2,desc=desc,un=int(un),preco=money(preco)))
    return rows

def load_needs():
    wb=load_workbook(RECON,data_only=True); ws=wb['Reconciliação']
    needs={}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        sku,necess,preco,saldo,emit_matriz,rest_filial,saldo_filial,emit_filial,falta,base_m,base_f=r
        needs[sku]=dict(total=int(necess), matriz=int(emit_matriz), filial=int(emit_filial), preco=money(preco))
    return needs

def allocate():
    rows=load_rateio_rows(); needs=load_needs(); alloc=[]
    for sku,n in needs.items():
        sku_rows=[x for x in rows if x['sku']==sku]
        if sum(x['un'] for x in sku_rows)!=n['total']:
            raise RuntimeError(f'Total mismatch {sku}')
        filial_total=n['filial']
        if filial_total==0:
            for x in sku_rows: alloc.append({**x,'estab':'matriz','un_emit':x['un']})
        else:
            floors=[]; rems=[]; allocated=0
            for i,x in enumerate(sku_rows):
                exact=D(x['un'])*D(filial_total)/D(n['total'])
                fl=int(exact.to_integral_value(rounding='ROUND_FLOOR'))
                floors.append(fl); allocated+=fl; rems.append((exact-D(fl), i))
            for _,i in sorted(rems, reverse=True)[:filial_total-allocated]: floors[i]+=1
            for x,fq in zip(sku_rows,floors):
                mq=x['un']-fq
                if mq: alloc.append({**x,'estab':'matriz','un_emit':mq})
                if fq: alloc.append({**x,'estab':'filial','un_emit':fq})
    return alloc

def load_templates():
    matriz=json.loads(MATRIZ_PREV.read_text())['payloads']
    templates={('matriz','07194128000182'): matriz['07194128000182'], ('matriz','45200547000179'): matriz['45200547000179'], ('matriz','63922116000106'): matriz['63922116000106']}
    for dest,file in DEST_FILE.items():
        p=json.loads((FILIAL_PREV_DIR/file).read_text())
        doc=p['contato']['numeroDocumento']
        templates[('filial',doc)]=p
    return templates

def item_maps(templates):
    maps={}
    for key,p in templates.items():
        m={}
        for it in p['itens']:
            sku=it['codigo']
            canonical=sku
            if sku=='YW1050': canonical='YW1050RC'
            if sku=='YW1520': canonical='YW1520RC'
            if sku=='YW640': canonical='YW640RC'
            m[canonical]=it
        maps[key]=m
    return maps

def build_payloads():
    alloc=allocate(); templates=load_templates(); maps=item_maps(templates)
    grouped=defaultdict(list)
    for a in alloc:
        grouped[(a['estab'], a['dest'], a['cnpj'].replace('.','').replace('/','').replace('-',''))].append(a)
    payloads=[]
    for (estab,dest,doc), items in sorted(grouped.items()):
        tmpl=json.loads(json.dumps(templates[(estab,doc)]))
        tmpl['dataOperacao']=brt_now()
        tmpl['naturezaOperacao']={'id': ACCOUNTS[estab]['natureza']}
        out_items=[]
        for a in sorted(items, key=lambda x:x['sku']):
            src=maps[(estab,doc)].get(a['sku'])
            if not src:
                raise RuntimeError(f'Item template missing {estab} {dest} {a["sku"]}')
            it={k:src[k] for k in ['codigo','descricao','unidade','classificacaoFiscal','origem'] if k in src}
            it['quantidade']=float(a['un_emit'])
            it['valor']=float(a['preco'])
            out_items.append(it)
        tmpl['itens']=out_items
        units=sum(D(i['quantidade']) for i in out_items)
        base=sum(D(i['quantidade'])*D(i['valor']) for i in out_items).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        payloads.append({'estab':estab,'dest':dest,'doc':doc,'payload':tmpl,'unidades':int(units),'base':float(base),'itens':len(out_items)})
    return payloads

def token(acct):
    return json.loads(ACCOUNTS[acct]['token_path'].read_text())['access_token']

def api(acct, method, path, data=None, timeout=30):
    url='https://www.bling.com.br/Api/v3'+path
    body=None
    headers={'Authorization':'Bearer '+token(acct),'Accept':'application/json'}
    if data is not None:
        body=json.dumps(data,ensure_ascii=False).encode()
        headers['Content-Type']='application/json'
    req=urllib.request.Request(url, data=body, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req,timeout=timeout) as resp:
            txt=resp.read().decode()
            parsed=json.loads(txt) if txt else {}
            return {'ok':True,'status':resp.status,'body':parsed}
    except urllib.error.HTTPError as e:
        txt=e.read().decode(errors='replace')
        try: parsed=json.loads(txt)
        except Exception: parsed=txt[:1000]
        return {'ok':False,'status':e.code,'body':parsed}

def create_notes(payloads):
    results=[]
    for p in payloads:
        r=api(p['estab'],'POST','/nfe',p['payload'])
        rec={k:p[k] for k in ['estab','dest','doc','unidades','base','itens']}
        rec.update({'create_status':r['status'],'create_ok':r['ok'],'create_body':r['body']})
        if r['ok']:
            data=r['body'].get('data') if isinstance(r['body'],dict) else None
            rec['id']=(data or {}).get('id')
            if rec.get('id'):
                g=api(p['estab'],'GET',f'/nfe/{rec["id"]}')
                rec['get_after_create']=g
        results.append(rec)
        time.sleep(1)
    return results

def send_notes(results):
    out=[]
    for rec in results:
        if not rec.get('id'):
            out.append(rec); continue
        s=api(rec['estab'],'POST',f'/nfe/{rec["id"]}/enviar')
        rec['send_status']=s['status']; rec['send_ok']=s['ok']; rec['send_body']=s['body']
        time.sleep(3)
        g=api(rec['estab'],'GET',f'/nfe/{rec["id"]}')
        rec['latest']=g
        out.append(rec)
    return out

def download_pdfs(results):
    for rec in results:
        body=(((rec.get('latest') or {}).get('body') or {}).get('data') or {}) if isinstance((rec.get('latest') or {}).get('body'),dict) else {}
        numero=body.get('numero') or str(rec.get('id'))
        link=body.get('linkPDF') or body.get('linkDanfe')
        rec['numero']=numero
        rec['situacao']=body.get('situacao')
        rec['chaveAcesso']=body.get('chaveAcesso')
        rec['valorNota']=body.get('valorNota')
        if link:
            try:
                pdf=urllib.request.urlopen(link,timeout=30).read()
                fp=OUTDIR/f"{rec['estab'].upper()}-NF-{numero}-{rec['dest'].replace(' ','_')}.pdf"
                fp.write_bytes(pdf)
                rec['pdf_path']=str(fp)
            except Exception as ex:
                rec['pdf_error']=str(ex)[:200]
    return results

def write_summary(payloads, results=None):
    summary={'generated_at_brt':brt_now(),'payloads_preview':[{k:p[k] for k in ['estab','dest','doc','unidades','base','itens']} for p in payloads]}
    if results is not None: summary['results']=results
    (OUTDIR/'manifest.json').write_text(json.dumps(summary,indent=2,ensure_ascii=False))
    with (OUTDIR/'summary.csv').open('w',newline='') as f:
        w=csv.writer(f); w.writerow(['estab','dest','doc','unidades','base','itens','id','numero','situacao','chave','valorNota','pdf_path'])
        rows=results if results is not None else payloads
        for r in rows:
            w.writerow([r.get('estab'),r.get('dest'),r.get('doc'),r.get('unidades'),r.get('base'),r.get('itens'),r.get('id'),r.get('numero'),r.get('situacao'),r.get('chaveAcesso'),r.get('valorNota'),r.get('pdf_path')])

if __name__=='__main__':
    mode=sys.argv[1] if len(sys.argv)>1 else 'preview'
    payloads=build_payloads()
    write_summary(payloads)
    print(json.dumps({'mode':mode,'count':len(payloads),'total_unidades':sum(p['unidades'] for p in payloads),'total_base':round(sum(p['base'] for p in payloads),2),'preview':[{k:p[k] for k in ['estab','dest','unidades','base','itens']} for p in payloads]},ensure_ascii=False,indent=2))
    if mode=='create':
        res=create_notes(payloads); write_summary(payloads,res); print(json.dumps(res,ensure_ascii=False,indent=2)[:6000])
    elif mode=='send':
        res=create_notes(payloads)
        if not all(r.get('id') for r in res):
            write_summary(payloads,res); print('CREATE_FAILED'); sys.exit(2)
        res=send_notes(res); res=download_pdfs(res); write_summary(payloads,res)
        print(json.dumps([{k:r.get(k) for k in ['estab','dest','id','numero','situacao','send_status','send_ok','chaveAcesso','valorNota','pdf_path','pdf_error']} for r in res],ensure_ascii=False,indent=2))
