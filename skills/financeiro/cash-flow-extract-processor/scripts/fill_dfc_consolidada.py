"""Preenche a DFC DIÁRIO REALIZADO - 2026.xlsx (consolidada anual).

Regras críticas (Knowledge File v10.1):
- Coluna = BASE_MES + dia
- SEMPRE validar via validate_column.py antes
- Linhas mapeadas em MAPA_CNPJ_CONS (CNPJ → Consolidada)
- Saldo inicial (R70) = saldo do GRUPO INTEIRO encadeado dia a dia
  (NÃO é a soma só das empresas que se moveram no dia)
- Nunca tocar em LINHAS_FORMULA_CONS
"""

import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from _constants import (
    D, BASES_MESES_2026, MAPA_CNPJ_CONS, LINHAS_FORMULA_CONS,
    LINHAS_ENT_OP, LINHAS_ENT_INV, LINHAS_ENT_FIN,
    LINHAS_SAI_OP, LINHAS_SAI_INV, LINHAS_SAI_FIN,
)
from validate_column import validar_coluna


def variacao_dia_empresa(lancamentos_da_empresa_no_dia: list) -> Decimal:
    """Calcula variação líquida de caixa do dia para uma empresa."""
    v = D("0")
    for lanc in lancamentos_da_empresa_no_dia:
        cl = lanc.get("classificacao") or {}
        linha = cl.get("linha_cnpj")
        if linha is None:
            continue
        valor = D(lanc["valor"])
        if linha == 69:
            v += valor if lanc["sinal"] == "+" else -valor
        elif linha in (LINHAS_ENT_OP | LINHAS_ENT_INV | LINHAS_ENT_FIN):
            v += valor
        elif linha in (LINHAS_SAI_OP | LINHAS_SAI_INV | LINHAS_SAI_FIN):
            v -= valor
    return v


def saldo_inicial_consolidado_por_dia(
    extratos_processados: list,
    saldos_30_anterior: dict[str, Decimal],
    dias_ordenados: list[str],
) -> dict[str, Decimal]:
    """R70 de cada dia = saldo do GRUPO INTEIRO encadeado.

    Args:
        saldos_30_anterior: {empresa: saldo no último dia do mês anterior}
        dias_ordenados: lista de dias (ISO) em ordem cronológica

    Returns:
        {dia: saldo_inicial_grupo}
    """
    saldo_grupo = sum(saldos_30_anterior.values(), D("0"))
    cons_ini = {}
    # Indexar lançamentos por (empresa, dia)
    lancs_por_emp_dia: dict[tuple[str, str], list] = defaultdict(list)
    for extrato in extratos_processados:
        if "erro" in extrato:
            continue
        emp = extrato["empresa"]
        for lanc in extrato.get("lancamentos", []):
            lancs_por_emp_dia[(emp, lanc["data"])].append(lanc)

    for dia in dias_ordenados:
        cons_ini[dia] = saldo_grupo
        var_dia = D("0")
        for empresa in saldos_30_anterior:
            lancs = lancs_por_emp_dia.get((empresa, dia), [])
            var_dia += variacao_dia_empresa(lancs)
        saldo_grupo += var_dia  # saldo fim deste dia = ini do próximo
    return cons_ini


def assert_linha_editavel_cons(linha: int):
    if linha in LINHAS_FORMULA_CONS:
        raise AssertionError(f"Linha {linha} é fórmula na consolidada — não preencher")


def preencher_dia_consolidada(
    planilha_path: Path,
    mes: str,
    dia: int,
    valores_por_linha: dict[int, Decimal],
    saldo_inicial_grupo: Decimal,
):
    """Preenche a coluna de 1 dia na DFC DIARIO REALIZADO.

    Args:
        planilha_path: caminho DFC DIARIO REALIZADO - 2026.xlsx
        mes: nome do mês em MAIÚSCULAS
        dia: dia do mês (1-31)
        valores_por_linha: {linha_consolidada: valor_consolidado_8_empresas}
        saldo_inicial_grupo: R70 (grupo inteiro encadeado)
    """
    col, _ = validar_coluna(planilha_path, mes, dia)  # raise se mês/dia não bater

    wb = openpyxl.load_workbook(planilha_path)
    ws = wb["DFC"] if "DFC" in wb.sheetnames else wb.active

    # Saldo inicial (R70)
    assert_linha_editavel_cons(70)
    ws.cell(row=70, column=col, value=float(saldo_inicial_grupo))

    for linha, valor in valores_por_linha.items():
        assert_linha_editavel_cons(linha)
        ws.cell(row=linha, column=col, value=float(valor))

    wb.save(planilha_path)


def consolidar_8_empresas(extratos_processados: list) -> dict:
    """Soma os valores das 8 empresas por (dia, linha_consolidada).

    Returns: {dia: {linha_cons: valor}}
    """
    por_dia: dict[str, dict[int, Decimal]] = defaultdict(lambda: defaultdict(lambda: D("0")))
    for extrato in extratos_processados:
        if "erro" in extrato:
            continue
        for lanc in extrato.get("lancamentos", []):
            cl = lanc.get("classificacao") or {}
            linha_cnpj = cl.get("linha_cnpj")
            if linha_cnpj is None or linha_cnpj not in MAPA_CNPJ_CONS:
                continue
            linha_cons = MAPA_CNPJ_CONS[linha_cnpj]
            valor = D(lanc["valor"])
            if linha_cnpj == 69:
                valor = valor if lanc["sinal"] == "+" else -valor
            por_dia[lanc["data"]][linha_cons] += valor
    return por_dia


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: ver docstring — este módulo é usado por script orquestrador")
        sys.exit(1)
