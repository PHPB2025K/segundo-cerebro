#!/usr/bin/env python3
"""
Shopee Extrato Multi-Conta — GB Importadora / Budamix
Gera extrato financeiro de uma ou todas as contas Shopee.
Cada conta vira uma aba no Excel + aba consolidada quando multi-conta.
"""
import argparse
import asyncio
import hashlib
import hmac
import json
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────
PARTNER_ID = 2031533
PARTNER_KEY = 'shpk494d456e49555973537464767779694476566c796e5579724648436b676c'
HOST = 'https://partner.shopeemobile.com'
TOKENS_DIR = Path('/root/.openclaw/workspace/integrations/shopee')
ACCOUNTS_FILE = TOKENS_DIR / 'accounts.json'
REPORTS_DIR = Path('/root/.openclaw/workspace/reports')
REPORTS_DIR.mkdir(exist_ok=True)

# Colors
GB_GREEN = "174C4C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
SUMMARY_GREEN = "E8F5E9"
SUMMARY_RED = "FFEBEE"
SUMMARY_YELLOW = "FFF8E1"
SUMMARY_BLUE = "E3F2FD"
CONSOL_HEADER = "1B5E5E"
CONSOL_SECTION = "2D8B8B"

# Rate limiter per-account
request_times = []

# 23 columns matching original format
COLUMNS = [
    'Ref. Externa', 'ID Origem', 'ID Usuário', 'Tipo Pagamento', 'Método',
    'Site', 'Tipo (técnico)', 'Classificação', 'Valor (R$)', 'Moeda',
    'Data Transação', 'Taxas (R$)', 'Valor Líquido (R$)', 'Moeda Liq.',
    'Data Liquidação', 'Status Liquidação', 'Valor Real (R$)', 'Cupom (R$)',
    'Metadata', 'ID Pedido', 'ID Envio', 'Modo Envio', 'ID Pack'
]
COL_WIDTHS = [18, 18, 12, 14, 12, 10, 16, 20, 14, 8, 20, 14, 14, 8, 20, 14, 14, 12, 30, 18, 18, 16, 12]
CURRENCY_COLS = {9, 12, 13, 17, 18}  # 1-indexed


# ── Account / Token Management ─────────────────────────────────────
def load_accounts():
    with open(ACCOUNTS_FILE) as f:
        data = json.load(f)
    accounts = {}
    for a in data['accounts']:
        alias = a['alias']
        display = a.get('shop_name', alias.replace('-', ' ').title())
        if alias == 'budamix-store':
            display = 'Budamix Store'
        accounts[alias] = {
            'alias': alias,
            'display': display,
            'shop_id': a['shop_id'],
            'token_file': TOKENS_DIR / a['token_file'],
        }
    return accounts


def load_tokens(account):
    with open(account['token_file']) as f:
        return json.load(f)


def save_tokens(account, tokens):
    with open(account['token_file'], 'w') as f:
        json.dump(tokens, f, indent=2)


def compute_sign(path, ts, access_token=None, shop_id=None):
    if access_token and shop_id:
        base = f"{PARTNER_ID}{path}{ts}{access_token}{shop_id}"
    else:
        base = f"{PARTNER_ID}{path}{ts}"
    return hmac.new(PARTNER_KEY.encode(), base.encode(), hashlib.sha256).hexdigest()


async def refresh_token(session, account, tokens):
    """Refresh access token for an account."""
    path = "/api/v2/auth/access_token/get"
    ts = int(time.time())
    sign = compute_sign(path, ts)  # public-level — no access_token/shop_id
    url = f"{HOST}{path}?partner_id={PARTNER_ID}&timestamp={ts}&sign={sign}"
    body = {
        "refresh_token": tokens['refresh_token'],
        "partner_id": PARTNER_ID,
        "shop_id": account['shop_id'],
    }
    async with session.post(url, json=body) as resp:
        data = await resp.json()
    if data.get('error'):
        raise RuntimeError(f"Token refresh failed for {account['alias']}: {data}")
    new_tokens = {
        'shop_id': account['shop_id'],
        'access_token': data['access_token'],
        'refresh_token': data['refresh_token'],
        'access_token_expire_in': data.get('expire_in', 14400),
        'refresh_token_expire_in': 2592000,
        'token_obtained_at': int(time.time()),
        'partner_id': PARTNER_ID,
        'updated_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
    }
    save_tokens(account, new_tokens)
    print(f"  ✓ Token refreshed for {account['alias']}")
    return new_tokens


async def ensure_valid_token(session, account):
    """Load tokens, refresh if expired, return valid tokens."""
    tokens = load_tokens(account)
    obtained = tokens.get('token_obtained_at', 0)
    expire_in = tokens.get('access_token_expire_in', 0)
    if time.time() > obtained + expire_in - 300:  # 5 min buffer
        print(f"  Token expirado para {account['alias']}, refreshing...")
        tokens = await refresh_token(session, account, tokens)
    return tokens


# ── API Helpers ─────────────────────────────────────────────────────
class ShopeeClient:
    """Per-account Shopee API client with rate limiting."""

    def __init__(self, account, tokens):
        self.account = account
        self.access_token = tokens['access_token']
        self.shop_id = tokens['shop_id']
        self.semaphore = asyncio.Semaphore(8)
        self.req_times = []

    def build_url(self, path, params=None):
        ts = int(time.time())
        sign = compute_sign(path, ts, self.access_token, self.shop_id)
        qs = (f"partner_id={PARTNER_ID}&shop_id={self.shop_id}"
              f"&timestamp={ts}&access_token={self.access_token}&sign={sign}")
        for k, v in (params or {}).items():
            qs += f"&{k}={v}"
        return f"{HOST}{path}?{qs}"

    async def rate_limit(self):
        now = time.time()
        self.req_times.append(now)
        self.req_times = [t for t in self.req_times if t > now - 1]
        if len(self.req_times) > 10:
            await asyncio.sleep(0.15)

    async def get(self, http_session, path, params=None):
        async with self.semaphore:
            await self.rate_limit()
            url = self.build_url(path, params)
            for attempt in range(3):
                try:
                    async with http_session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                        data = await resp.json()
                        if data.get('error') and 'too many' in str(data.get('message', '')).lower():
                            await asyncio.sleep(2 ** attempt)
                            # Rebuild URL with fresh timestamp
                            url = self.build_url(path, params)
                            continue
                        return data
                except Exception as e:
                    if attempt < 2:
                        await asyncio.sleep(1)
                        url = self.build_url(path, params)
                    else:
                        return {"error": str(e)}
            return {"error": "max retries"}


# ── Data Collection ─────────────────────────────────────────────────
async def get_all_orders(client, http_session, time_from, time_to):
    all_orders = []
    windows = []
    current = time_from
    while current < time_to:
        window_end = min(current + 15 * 86400, time_to)
        windows.append((current, window_end))
        current = window_end

    for status in ['COMPLETED', 'CANCELLED']:
        for wf, wt in windows:
            cursor = "0"
            while True:
                r = await client.get(http_session, '/api/v2/order/get_order_list', {
                    'time_range_field': 'create_time',
                    'time_from': wf,
                    'time_to': wt,
                    'page_size': 50,
                    'cursor': cursor,
                    'order_status': status,
                })
                resp = r.get('response', {})
                for o in resp.get('order_list', []):
                    all_orders.append({'order_sn': o['order_sn'], 'status': status})
                if not resp.get('more'):
                    break
                cursor = resp.get('next_cursor', '0')

    seen = set()
    unique = []
    for o in all_orders:
        if o['order_sn'] not in seen:
            seen.add(o['order_sn'])
            unique.append(o)
    return unique


async def get_order_details_batch(client, http_session, order_sns):
    details = {}
    for i in range(0, len(order_sns), 50):
        batch = order_sns[i:i + 50]
        r = await client.get(http_session, '/api/v2/order/get_order_detail', {
            'order_sn_list': ','.join(batch),
            'response_optional_fields': 'buyer_user_id,buyer_username,estimated_shipping_fee,actual_shipping_fee,pay_time,ship_by_date,tracking_no,shipping_carrier',
        })
        for order in r.get('response', {}).get('order_list', []):
            details[order['order_sn']] = order
        print(f"    Detail: {min(i + 50, len(order_sns))}/{len(order_sns)}", flush=True)
    return details


async def get_all_escrows(client, http_session, order_sns):
    results = {}
    total = len(order_sns)
    done = [0]
    errors = [0]

    async def fetch_one(sn):
        r = await client.get(http_session, '/api/v2/payment/get_escrow_detail', {'order_sn': sn})
        done[0] += 1
        if done[0] % 50 == 0 or done[0] == total:
            print(f"    Escrow: {done[0]}/{total} (erros: {errors[0]})", flush=True)
        if r.get('error') and r['error'] != '':
            errors[0] += 1
            return
        results[sn] = r.get('response', {})

    for i in range(0, total, 50):
        chunk = order_sns[i:i + 50]
        await asyncio.gather(*(fetch_one(sn) for sn in chunk))

    return results, errors[0]


# ── Row / Summary Builders ──────────────────────────────────────────
def ts_to_datetime(ts):
    if not ts:
        return None
    dt = datetime.fromtimestamp(ts, tz=timezone(timedelta(hours=-3)))
    return dt.replace(tzinfo=None)


def fmt_dt(dt):
    return dt.strftime('%d/%m/%Y %H:%M:%S') if dt else ''


def classify_order(status):
    m = {
        'COMPLETED': ('SETTLEMENT', 'Venda Concluída'),
        'CANCELLED': ('REFUND', 'Devolução/Cancelamento'),
    }
    return m.get(status, ('OTHER', status))


def liq_status(order_status, escrow):
    if order_status == 'COMPLETED':
        amt = escrow.get('order_income', {}).get('escrow_amount', 0)
        return 'Recebido' if amt > 0 else 'Pendente'
    if order_status == 'CANCELLED':
        return 'Cancelado'
    return 'Pendente'


def build_rows_and_stats(order_sns, status_map, order_details, escrow_details, shop_id):
    """Build data rows and stats dict from raw API data."""
    rows = []
    stats = {
        'count_completed': 0,
        'count_cancelled': 0,
        'total_bruto': 0,
        'total_liquido': 0,
        'total_taxas': 0,
        'total_cupons': 0,
        'total_frete': 0,
        'method_count': defaultdict(int),
        'method_bruto': defaultdict(float),
        'method_liquido': defaultdict(float),
        'products': defaultdict(lambda: {'qty': 0, 'bruto': 0, 'name': '', 'sku': ''}),
    }

    for sn in order_sns:
        order = order_details.get(sn, {})
        escrow = escrow_details.get(sn)
        order_status = status_map[sn]
        tipo_tec, classificacao = classify_order(order_status)
        create_time = ts_to_datetime(order.get('create_time', 0))

        if escrow is None:
            rows.append([
                sn, sn, shop_id, '', '', 'SHOPEE_BR',
                tipo_tec, classificacao, 0, 'BRL',
                fmt_dt(create_time), 0, 0, 'BRL',
                '', 'Sem dados', 0, 0,
                '{"error":"no escrow"}', sn, '', '', ''
            ])
            if order_status == 'CANCELLED':
                stats['count_cancelled'] += 1
            continue

        income = escrow.get('order_income', {})
        buyer_info = escrow.get('buyer_payment_info', {})

        bruto = income.get('cost_of_goods_sold', 0) or 0
        liquido = income.get('escrow_amount', 0) or 0
        commission = income.get('commission_fee', 0) or 0
        service = income.get('service_fee', 0) or 0
        transaction = income.get('credit_card_transaction_fee', 0) or 0
        campaign = income.get('campaign_fee', 0) or 0
        fbs = income.get('fbs_fee', 0) or 0
        final_shipping = income.get('final_shipping_fee', 0) or 0
        buyer_shipping = income.get('buyer_paid_shipping_fee', 0) or 0
        seller_voucher = abs(income.get('seller_voucher', 0) or 0)
        shopee_voucher = abs(income.get('shopee_voucher', 0) or 0)
        coins = abs(income.get('coins', 0) or 0)

        taxas = commission + service + transaction + campaign + fbs
        cupons = seller_voucher + shopee_voucher + coins
        method = buyer_info.get('buyer_payment_method', '')
        pay_time = ts_to_datetime(order.get('pay_time', 0))
        tracking = order.get('tracking_no', '') or ''
        carrier = order.get('shipping_carrier', '') or ''

        metadata = json.dumps({
            'commission_fee': commission, 'service_fee': service,
            'transaction_fee': transaction, 'campaign_fee': campaign,
            'fbs_fee': fbs, 'final_shipping_fee': final_shipping,
            'buyer_paid_shipping_fee': buyer_shipping,
            'actual_shipping_fee': income.get('actual_shipping_fee', 0),
            'seller_voucher': seller_voucher, 'shopee_voucher': shopee_voucher,
            'coins': coins,
            'escrow_amount_after_adjustment': income.get('escrow_amount_after_adjustment', 0),
        }, ensure_ascii=False)

        rows.append([
            sn, sn, shop_id, method, method, 'SHOPEE_BR',
            tipo_tec, classificacao, round(bruto, 2), 'BRL',
            fmt_dt(create_time), round(-taxas, 2), round(liquido, 2), 'BRL',
            fmt_dt(pay_time), liq_status(order_status, escrow),
            round(liquido, 2), round(-cupons, 2) if cupons > 0 else 0,
            metadata, sn, tracking, carrier, ''
        ])

        stats['method_count'][method] += 1
        if order_status == 'COMPLETED':
            stats['count_completed'] += 1
            stats['total_bruto'] += bruto
            stats['total_liquido'] += liquido
            stats['total_taxas'] += taxas
            stats['total_cupons'] += cupons
            stats['total_frete'] += final_shipping
            stats['method_bruto'][method] += bruto
            stats['method_liquido'][method] += liquido
            # products
            for item in order.get('item_list', []):
                key = item.get('item_sku', '') or item.get('item_id', '')
                p = stats['products'][key]
                p['qty'] += item.get('model_quantity_purchased', 0) or 1
                p['bruto'] += bruto  # approximate per-order
                p['name'] = item.get('item_name', p['name'] or '')
                p['sku'] = item.get('item_sku', p['sku'] or '')
        elif order_status == 'CANCELLED':
            stats['count_cancelled'] += 1

    rows.sort(key=lambda r: r[10] if r[10] else '', reverse=True)
    return rows, stats


def build_summary(stats):
    tb = stats['total_bruto']
    pct = lambda v: f"{(v / tb * 100):.1f}%" if tb > 0 else "0%"
    summary = [
        {'label': 'Vendas Concluídas', 'qty': stats['count_completed'], 'value': round(tb, 2),
         'pct': '100%', 'desc': 'Total bruto de vendas concluídas', 'color': 'green'},
        {'label': 'Comissões + Taxas', 'qty': stats['count_completed'], 'value': round(-stats['total_taxas'], 2),
         'pct': pct(stats['total_taxas']), 'desc': 'Comissão + serviço + transação + campanha + FBS', 'color': 'red'},
        {'label': 'Cupons/Vouchers', 'qty': '', 'value': round(-stats['total_cupons'], 2),
         'pct': pct(stats['total_cupons']), 'desc': 'Voucher vendedor + Shopee + moedas', 'color': 'red'},
        {'label': 'Frete Líquido', 'qty': '', 'value': round(stats['total_frete'], 2),
         'pct': pct(abs(stats['total_frete'])), 'desc': 'Frete líquido', 'color': 'yellow'},
        {'label': 'Cancelamentos', 'qty': stats['count_cancelled'], 'value': 0,
         'pct': '', 'desc': 'Pedidos cancelados no período', 'color': 'red'},
        {'label': 'SALDO LÍQUIDO', 'qty': stats['count_completed'], 'value': round(stats['total_liquido'], 2),
         'pct': pct(stats['total_liquido']), 'desc': 'Valor efetivamente recebido', 'color': 'blue'},
    ]
    for method, count in sorted(stats['method_count'].items(), key=lambda x: -x[1]):
        bm = round(stats['method_bruto'].get(method, 0), 2)
        lm = round(stats['method_liquido'].get(method, 0), 2)
        summary.append({
            'label': f'  → {method}', 'qty': count, 'value': lm,
            'pct': pct(bm) if bm > 0 else '0%',
            'desc': f'Bruto R$ {bm:,.2f} • Líquido R$ {lm:,.2f}', 'color': 'green'
        })
    return summary


# ── Excel Builders ──────────────────────────────────────────────────
# Shared styles
_header_fill = PatternFill(start_color=GB_GREEN, end_color=GB_GREEN, fill_type="solid")
_header_font = Font(bold=True, color=WHITE, size=11)
_title_font = Font(bold=True, size=14, color=GB_GREEN)
_subtitle_font = Font(size=10, color="666666")
_zebra_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
_thin_border = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
)
_currency_fmt = '#,##0.00'
_pct_fmt = '0.0%'

COLOR_MAP = {
    'green': SUMMARY_GREEN, 'red': SUMMARY_RED,
    'yellow': SUMMARY_YELLOW, 'blue': SUMMARY_BLUE,
}


def write_account_sheet(ws, rows, summary, title, subtitle):
    """Write a single-account sheet (23 columns + summary)."""
    ws.merge_cells('A1:W1')
    c = ws['A1']
    c.value = title
    c.font = _title_font
    c.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:W2')
    c = ws['A2']
    c.value = subtitle
    c.font = _subtitle_font
    c.alignment = Alignment(horizontal='center')

    # Header row 4
    for ci, h in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = _header_fill
        cell.font = _header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = _thin_border

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data
    for ri, row_data in enumerate(rows, 5):
        fill = _zebra_fill if (ri - 5) % 2 == 1 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _thin_border
            if fill:
                cell.fill = fill
            if ci in CURRENCY_COLS and isinstance(val, (int, float)):
                cell.number_format = _currency_fmt
                cell.alignment = Alignment(horizontal='right')

    # Summary
    sr = len(rows) + 6
    ws.merge_cells(f'A{sr}:W{sr}')
    ws[f'A{sr}'].value = 'RESUMO FINANCEIRO'
    ws[f'A{sr}'].font = Font(bold=True, size=13, color=GB_GREEN)
    ws[f'A{sr}'].alignment = Alignment(horizontal='center')

    sr += 1
    for ci, h in enumerate(['Categoria', 'Qtd', 'Valor (R$)', '% do Bruto', 'Descrição'], 1):
        cell = ws.cell(row=sr, column=ci, value=h)
        cell.fill = _header_fill
        cell.font = _header_font
        cell.border = _thin_border

    for i, s in enumerate(summary):
        r = sr + 1 + i
        fill = PatternFill(start_color=COLOR_MAP.get(s['color'], WHITE),
                           end_color=COLOR_MAP.get(s['color'], WHITE), fill_type="solid")
        for ci, v in enumerate([s['label'], s['qty'], s['value'], s['pct'], s['desc']], 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.fill = fill
            cell.border = _thin_border
            if ci == 3 and isinstance(v, (int, float)):
                cell.number_format = _currency_fmt

    ws.freeze_panes = 'A5'


def write_consolidated_sheet(ws, all_account_stats, all_account_names):
    """Write the Consolidado tab."""
    consol_fill = PatternFill(start_color=CONSOL_HEADER, end_color=CONSOL_HEADER, fill_type="solid")
    section_fill = PatternFill(start_color=CONSOL_SECTION, end_color=CONSOL_SECTION, fill_type="solid")
    white_font = Font(bold=True, color=WHITE, size=11)
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)

    num_accounts = len(all_account_names)
    total_cols = num_accounts + 2  # Métrica + accounts + TOTAL

    ws.column_dimensions['A'].width = 28
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row = 1
    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    c = ws.cell(row=row, column=1, value='EXTRATO CONSOLIDADO — SHOPEE / BUDAMIX')
    c.font = Font(bold=True, size=14, color=GB_GREEN)
    c.alignment = Alignment(horizontal='center')
    row += 2

    # ── Section 1: Comparativo por Conta ────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    c = ws.cell(row=row, column=1, value='COMPARATIVO POR CONTA')
    c.fill = consol_fill
    c.font = white_font
    c.alignment = Alignment(horizontal='center')
    row += 1

    # Header
    headers = ['Métrica'] + all_account_names + ['TOTAL']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = section_fill
        cell.font = white_font
        cell.border = _thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Compute totals
    totals = {
        'completed': sum(s['count_completed'] for s in all_account_stats.values()),
        'cancelled': sum(s['count_cancelled'] for s in all_account_stats.values()),
        'bruto': sum(s['total_bruto'] for s in all_account_stats.values()),
        'taxas': sum(s['total_taxas'] for s in all_account_stats.values()),
        'cupons': sum(s['total_cupons'] for s in all_account_stats.values()),
        'frete': sum(s['total_frete'] for s in all_account_stats.values()),
        'liquido': sum(s['total_liquido'] for s in all_account_stats.values()),
    }

    def pct_taxas(s):
        return s['total_taxas'] / s['total_bruto'] if s['total_bruto'] > 0 else 0

    def margem(s):
        return s['total_liquido'] / s['total_bruto'] if s['total_bruto'] > 0 else 0

    def ticket_bruto(s):
        return s['total_bruto'] / s['count_completed'] if s['count_completed'] > 0 else 0

    def ticket_liq(s):
        return s['total_liquido'] / s['count_completed'] if s['count_completed'] > 0 else 0

    aliases = list(all_account_stats.keys())

    metrics = [
        ('Pedidos Concluídos', lambda s: s['count_completed'], 'int',
         totals['completed']),
        ('Pedidos Cancelados', lambda s: s['count_cancelled'], 'int',
         totals['cancelled']),
        ('Valor Bruto (R$)', lambda s: round(s['total_bruto'], 2), 'money',
         round(totals['bruto'], 2)),
        ('Comissões + Taxas (R$)', lambda s: round(s['total_taxas'], 2), 'money',
         round(totals['taxas'], 2)),
        ('% Taxas s/ Bruto', pct_taxas, 'pct',
         totals['taxas'] / totals['bruto'] if totals['bruto'] > 0 else 0),
        ('Cupons/Vouchers (R$)', lambda s: round(s['total_cupons'], 2), 'money',
         round(totals['cupons'], 2)),
        ('Frete Líquido (R$)', lambda s: round(s['total_frete'], 2), 'money',
         round(totals['frete'], 2)),
        ('Valor Líquido (R$)', lambda s: round(s['total_liquido'], 2), 'money',
         round(totals['liquido'], 2)),
        ('Margem Líquida', margem, 'pct',
         totals['liquido'] / totals['bruto'] if totals['bruto'] > 0 else 0),
        ('Ticket Médio Bruto', ticket_bruto, 'money',
         totals['bruto'] / totals['completed'] if totals['completed'] > 0 else 0),
        ('Ticket Médio Líquido', ticket_liq, 'money',
         totals['liquido'] / totals['completed'] if totals['completed'] > 0 else 0),
    ]

    for mi, (label, fn, fmt, total_val) in enumerate(metrics):
        fill = _zebra_fill if mi % 2 == 1 else None
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold_font
        cell.border = _thin_border
        if fill:
            cell.fill = fill

        for ai, alias in enumerate(aliases):
            s = all_account_stats[alias]
            val = fn(s)
            cell = ws.cell(row=row, column=2 + ai, value=val)
            cell.border = _thin_border
            if fill:
                cell.fill = fill
            if fmt == 'money':
                cell.number_format = _currency_fmt
            elif fmt == 'pct':
                cell.number_format = _pct_fmt
            cell.alignment = Alignment(horizontal='right')

        # TOTAL column
        cell = ws.cell(row=row, column=total_cols, value=total_val)
        cell.font = bold_font
        cell.border = _thin_border
        if fill:
            cell.fill = fill
        if fmt == 'money':
            cell.number_format = _currency_fmt
        elif fmt == 'pct':
            cell.number_format = _pct_fmt
        cell.alignment = Alignment(horizontal='right')
        row += 1

    row += 2

    # ── Section 2: Breakdown por Método de Pagamento ────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value='BREAKDOWN POR MÉTODO DE PAGAMENTO (CONSOLIDADO)')
    c.fill = consol_fill
    c.font = white_font
    c.alignment = Alignment(horizontal='center')
    row += 1

    pay_headers = ['Método', 'Qtd', 'Valor Bruto (R$)', 'Valor Líquido (R$)', '% do Total']
    for ci, h in enumerate(pay_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = section_fill
        cell.font = white_font
        cell.border = _thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Aggregate methods across accounts
    agg_method_count = defaultdict(int)
    agg_method_bruto = defaultdict(float)
    agg_method_liquido = defaultdict(float)
    for s in all_account_stats.values():
        for m, c in s['method_count'].items():
            agg_method_count[m] += c
        for m, v in s['method_bruto'].items():
            agg_method_bruto[m] += v
        for m, v in s['method_liquido'].items():
            agg_method_liquido[m] += v

    total_bruto_all = totals['bruto']
    for mi, (method, count) in enumerate(sorted(agg_method_count.items(), key=lambda x: -x[1])):
        fill = _zebra_fill if mi % 2 == 1 else None
        bruto_m = round(agg_method_bruto.get(method, 0), 2)
        liq_m = round(agg_method_liquido.get(method, 0), 2)
        pct_val = bruto_m / total_bruto_all if total_bruto_all > 0 else 0

        vals = [method, count, bruto_m, liq_m, pct_val]
        fmts = [None, None, _currency_fmt, _currency_fmt, _pct_fmt]
        for ci, (v, nf) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = _thin_border
            if fill:
                cell.fill = fill
            if nf:
                cell.number_format = nf
            if ci >= 2:
                cell.alignment = Alignment(horizontal='right')
        row += 1

    row += 2

    # ── Section 3: Top 10 Produtos ──────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value='TOP 10 PRODUTOS (POR VOLUME, CONSOLIDADO)')
    c.fill = consol_fill
    c.font = white_font
    c.alignment = Alignment(horizontal='center')
    row += 1

    prod_headers = ['#', 'Produto', 'SKU', 'Qtd Vendida', 'Valor Bruto (R$)', 'Conta']
    for ci, h in enumerate(prod_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = section_fill
        cell.font = white_font
        cell.border = _thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Aggregate products
    agg_products = defaultdict(lambda: {'qty': 0, 'bruto': 0, 'name': '', 'sku': '', 'accounts': set()})
    for alias, s in all_account_stats.items():
        display = all_account_stats[alias].get('_display', alias)
        for key, p in s['products'].items():
            ap = agg_products[key]
            ap['qty'] += p['qty']
            ap['bruto'] += p['bruto']
            if p['name']:
                ap['name'] = p['name']
            if p['sku']:
                ap['sku'] = p['sku']
            ap['accounts'].add(display)

    top10 = sorted(agg_products.values(), key=lambda x: -x['qty'])[:10]
    for pi, p in enumerate(top10):
        fill = _zebra_fill if pi % 2 == 1 else None
        accounts_str = ', '.join(sorted(p['accounts']))
        vals = [pi + 1, p['name'][:60], p['sku'], p['qty'], round(p['bruto'], 2), accounts_str]
        fmts = [None, None, None, None, _currency_fmt, None]
        for ci, (v, nf) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = _thin_border
            if fill:
                cell.fill = fill
            if nf:
                cell.number_format = nf
        row += 1

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['F'].width = 30


# ── Process Single Account ──────────────────────────────────────────
async def process_account(http_session, account, time_from, time_to):
    """Process a single Shopee account. Returns (rows, stats) or raises."""
    print(f"\n{'─' * 50}")
    print(f"Conta: {account['display']} ({account['alias']}) — Shop {account['shop_id']}")
    print(f"{'─' * 50}")

    tokens = await ensure_valid_token(http_session, account)
    client = ShopeeClient(account, tokens)

    print("  [1/4] Buscando pedidos...")
    orders = await get_all_orders(client, http_session, time_from, time_to)
    print(f"    → {len(orders)} pedidos encontrados")

    if not orders:
        empty_stats = {
            'count_completed': 0, 'count_cancelled': 0,
            'total_bruto': 0, 'total_liquido': 0, 'total_taxas': 0,
            'total_cupons': 0, 'total_frete': 0,
            'method_count': {}, 'method_bruto': {}, 'method_liquido': {},
            'products': {},
        }
        return [], empty_stats

    order_sns = [o['order_sn'] for o in orders]
    status_map = {o['order_sn']: o['status'] for o in orders}

    print("  [2/4] Buscando detalhes...")
    order_details = await get_order_details_batch(client, http_session, order_sns)

    print("  [3/4] Buscando escrow...")
    escrow_details, escrow_errors = await get_all_escrows(client, http_session, order_sns)

    print("  [4/4] Montando dados...")
    rows, stats = build_rows_and_stats(order_sns, status_map, order_details, escrow_details, account['shop_id'])
    stats['escrow_errors'] = escrow_errors

    print(f"  ✓ {stats['count_completed']} concluídos, {stats['count_cancelled']} cancelados, "
          f"bruto R$ {stats['total_bruto']:,.2f}, líquido R$ {stats['total_liquido']:,.2f}")
    return rows, stats


# ── Main ────────────────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser(description='Extrato Shopee Multi-Conta')
    parser.add_argument('--inicio', required=True, help='Data início (YYYY-MM-DD)')
    parser.add_argument('--fim', required=True, help='Data fim (YYYY-MM-DD)')
    parser.add_argument('--conta', default='todas',
                        help='Nome da conta ou "todas" (default: todas)')
    args = parser.parse_args()

    start_time = time.time()
    time_from = int(time.mktime(time.strptime(args.inicio, '%Y-%m-%d')))
    # fim = end of day
    fim_date = datetime.strptime(args.fim, '%Y-%m-%d')
    time_to = int(time.mktime((fim_date + timedelta(days=1) - timedelta(seconds=1)).timetuple()))

    period_start = datetime.strptime(args.inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
    period_end = datetime.strptime(args.fim, '%Y-%m-%d').strftime('%d/%m/%Y')

    accounts = load_accounts()

    if args.conta == 'todas':
        target_accounts = list(accounts.values())
    elif args.conta in accounts:
        target_accounts = [accounts[args.conta]]
    else:
        print(f"Conta '{args.conta}' não encontrada. Disponíveis: {', '.join(accounts.keys())}")
        sys.exit(1)

    multi = len(target_accounts) > 1

    print("=" * 60)
    print("Extrato Financeiro — Shopee / Budamix")
    print(f"Período: {period_start} a {period_end}")
    print(f"Contas: {', '.join(a['display'] for a in target_accounts)}")
    print("=" * 60)

    all_rows = {}     # alias -> rows
    all_stats = {}    # alias -> stats
    errors = {}       # alias -> error message
    account_names = []  # ordered display names

    async with aiohttp.ClientSession() as http_session:
        for account in target_accounts:
            alias = account['alias']
            account_names.append(account['display'])
            try:
                rows, stats = await process_account(http_session, account, time_from, time_to)
                stats['_display'] = account['display']
                all_rows[alias] = rows
                all_stats[alias] = stats
            except Exception as e:
                print(f"  ✗ ERRO na conta {alias}: {e}")
                errors[alias] = str(e)
                all_rows[alias] = []
                all_stats[alias] = {
                    'count_completed': 0, 'count_cancelled': 0,
                    'total_bruto': 0, 'total_liquido': 0, 'total_taxas': 0,
                    'total_cupons': 0, 'total_frete': 0,
                    'method_count': {}, 'method_bruto': {}, 'method_liquido': {},
                    'products': {}, '_display': account['display'],
                    'escrow_errors': 0,
                }

    # ── Build Excel ─────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print("Gerando Excel...")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for account in target_accounts:
        alias = account['alias']
        display = account['display']
        rows = all_rows[alias]
        stats = all_stats[alias]
        summary = build_summary(stats)

        ws = wb.create_sheet(title=display[:31])  # Excel max 31 chars
        title_str = f'Extrato Financeiro — Shopee / {display}'
        subtitle_str = (f'GB Importadora / Budamix • Shop ID {account["shop_id"]} '
                        f'• Período: {period_start} a {period_end}')
        write_account_sheet(ws, rows, summary, title_str, subtitle_str)

    if multi:
        ws_consol = wb.create_sheet(title='Consolidado')
        write_consolidated_sheet(ws_consol, all_stats, account_names)

    # File name
    if multi:
        filename = f"extrato-shopee-todas-{args.inicio}-{args.fim}.xlsx"
    else:
        filename = f"extrato-shopee-{target_accounts[0]['alias']}-{args.inicio}-{args.fim}.xlsx"
    output_path = REPORTS_DIR / filename
    wb.save(output_path)

    elapsed = time.time() - start_time

    # ── Report ──────────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print("RESULTADO FINAL")
    print(f"{'=' * 60}")
    print(f"Arquivo: {output_path}")
    print(f"Tempo: {elapsed:.1f}s")
    print()

    grand_bruto = 0
    grand_liquido = 0
    grand_taxas = 0
    grand_completed = 0
    grand_cancelled = 0

    for account in target_accounts:
        alias = account['alias']
        s = all_stats[alias]
        print(f"  {account['display']}:")
        print(f"    Pedidos: {s['count_completed']} concluídos, {s['count_cancelled']} cancelados")
        print(f"    Bruto:   R$ {s['total_bruto']:,.2f}")
        print(f"    Taxas:   R$ {s['total_taxas']:,.2f}")
        print(f"    Líquido: R$ {s['total_liquido']:,.2f}")
        esc_err = s.get('escrow_errors', 0)
        if esc_err:
            print(f"    Erros escrow: {esc_err}")
        grand_bruto += s['total_bruto']
        grand_liquido += s['total_liquido']
        grand_taxas += s['total_taxas']
        grand_completed += s['count_completed']
        grand_cancelled += s['count_cancelled']
        print()

    if multi:
        print(f"  CONSOLIDADO:")
        print(f"    Pedidos: {grand_completed} concluídos, {grand_cancelled} cancelados")
        print(f"    Bruto:   R$ {grand_bruto:,.2f}")
        print(f"    Taxas:   R$ {grand_taxas:,.2f}")
        print(f"    Líquido: R$ {grand_liquido:,.2f}")
        print()

    if errors:
        print(f"  ERROS:")
        for alias, err in errors.items():
            print(f"    {alias}: {err}")

    print(f"{'=' * 60}")


if __name__ == '__main__':
    asyncio.run(main())
