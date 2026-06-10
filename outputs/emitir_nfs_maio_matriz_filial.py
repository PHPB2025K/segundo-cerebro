#!/usr/bin/env python3
import json, csv, urllib.request, urllib.error, sys, time
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from openpyxl import load_workbook

ROOT=Path('/root/segundo-cerebro')
OUTDIR=ROOT/'outputs'/'nfs-maio-2026-matriz-filial-final'
RATEIO=ROOT/'outputs'/'rateio-fiscal-maio-potes-cnpjs-simples-2026-06-08.xlsx'
RECON=ROOT/'outputs'/'reconciliacao-matriz-filial-maio-2026'/'reconciliacao-matriz-filial-maio-2026-refinada.xlsx'
MATRIZ_PREV=ROOT/'outputs'/'bling-nfe-matriz-maio-cnpjs-simples-2026-06-09.json'
FILIAL_PREV_DIR=ROOT/'tmp'/'bling_payloads_abril'

ACCOUNTS={
 'matriz': {'token_path': Path('/root/.openclaw/workspace/scripts/bling-oauth/tokens-matriz.json'),'natureza':15106994870,'label':'Matriz'},
 'filial': {'token_path': Path('/root/.openclaw/workspace/scripts/bling-oauth/tokens-filial.json'),'natureza':15107403191,'label':'Filial'},
}
DEST_FILE={'GB Comércio':'GB_Comercio.json','Trades':'Trades.json','Broglio':'Broglio.json'}

def nowstr(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
def D(x): return Decimal(str(x))
def money(x): return D(x).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def load_rateio_rows():
    wb=load_workbook(RATEIO,data_only=True); ws=wb['Detalhe por SKU']
    rows=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        dest,cnpj,grupo,sku,desc,un,custo,preco,valor,icms=r
        rows.append(dict(dest=dest,cnpj=''.join(ch for ch in str(cnpj) if ch.isdigit()),sku='YW320SQ' if sku=='YW320' else sku,desc=desc,un=int(un),preco=money(preco)))
    return rows

def load_needs():
    wb=load_workbook(RECON,data_only=True); ws=wb['Reconciliação']
    needs={}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        sku,necess,preco,saldo,emit_matriz,rest_filial,saldo_filial,emit_filial,falta,base_m,base_f=r
        needs[sku]=dict(total=int(necess),matriz=int(emit_matriz),filial=int(emit_filial))
    return needs

def allocate():
    rows=load_rateio_rows(); needs=load_needs(); alloc=[]
    for sku,n in needs.items():
        sku_rows=[x for x in rows if x['sku']==sku]
        assert sum(x['un'] for x in sku_rows)==n['total'], (sku, sum(x['un'] for x in sku_rows), n['total'])
        filial_total=n['filial']
        if filial_total==0:
            for x in sku_rows: alloc.append({**x,'estab':'matriz','un_emit':x['un']})
            continue
        floors=[]; rems=[]; allocated=0
        for i,x in enumerate(sku_rows):
            exact=D(x['un'])*D(filial_total)/D(n['total'])
            fl=int(exact.to_integral_value(rounding='ROUND_FLOOR'))
            floors.append(fl); rems.append((exact-D(fl),i)); allocated+=fl
        for _,i in sorted(rems, reverse=True)[:filial_total-allocated]: floors[i]+=1
        for x,fq in zip(sku_rows,floors):
            mq=x['un']-fq
            if mq: alloc.append({**x,'estab':'matriz','un_emit':mq})
            if fq: alloc.append({**x,'estab':'filial','un_emit':fq})
    return alloc

def load_templates():
    templates={}
    matriz=json.loads(MATRIZ_PREV.read_text())['payloads']
    for doc,p in matriz.items(): templates[('matriz',doc)]=p
    for dest,file in DEST_FILE.items():
        p=json.loads((FILIAL_PREV_DIR/file).read_text())
        templates[('filial',p['contato']['numeroDocumento'])]=p
    return templates

def maps_for_templates(templates):
    out={}
    for key,p in templates.items():
        m={}
        for it in p['itens']:
            sku=it['codigo']
            if sku=='YW1050': sku='YW1050RC'
            elif sku=='YW1520': sku='YW1520RC'
            elif sku=='YW640': sku='YW640RC'
            m[sku]=it
        out[key]=m
    return out

def build_payloads():
    alloc=allocate(); templates=load_templates(); itemmaps=maps_for_templates(templates)
    grouped=defaultdict(list)
    for a in alloc: grouped[(a['estab'],a['dest'],a['cnpj'])].append(a)
    payloads=[]
    for (estab,dest,doc), items in sorted(grouped.items()):
        tmpl=json.loads(json.dumps(templates[(estab,doc)]))
        tmpl['dataOperacao']=nowstr()
        tmpl['naturezaOperacao']={'id': ACCOUNTS[estab]['natureza']}
        out_items=[]
        for a in sorted(items, key=lambda x:x['sku']):
            src=itemmaps[(estab,doc)][a['sku']]
            item={k:src[k] for k in ['codigo','descricao','unidade','classificacaoFiscal','origem'] if k in src}
            item['quantidade']=float(a['un_emit'])
            item['valor']=float(a['preco'])
            out_items.append(item)
        tmpl['itens']=out_items
        unidades=sum(int(i['quantidade']) for i in out_items)
        base=float(sum((D(i['quantidade'])*D(i['valor']) for i in out_items), Decimal('0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        payloads.append({'estab':estab,'dest':dest,'doc':doc,'payload':tmpl,'unidades':unidades,'base':base,'itens':len(out_items)})
    return payloads

def token(acct): return json.loads(ACCOUNTS[acct]['token_path'].read_text())['access_token']

def api(acct, method, path, data=None, timeout=30):
    req=urllib.request.Request('https://www.bling.com.br/Api/v3'+path, method=method, headers={'Authorization':'Bearer '+token(acct),'Accept':'application/json'})
    if data is not None:
        body=json.dumps(data,ensure_ascii=False).encode()
        req.data=body
        req.add_header('Content-Type','application/json')
    try:
        with urllib.request.urlopen(req,timeout=timeout) as resp:
            txt=resp.read().decode()
            return {'ok':True,'status':resp.status,'body':json.loads(txt) if txt else {}}
    except urllib.error.HTTPError as e:
        txt=e.read().decode(errors='replace')
        try: body=json.loads(txt)
        except Exception: body=txt[:2000]
        return {'ok':False,'status':e.code,'body':body}

def write_summary(payloads, results=None):
    OUTDIR.mkdir(parents=True, exist_ok=True)
    data={'generated_at_brt':nowstr(),'payloads_preview':[{k:p[k] for k in ['estab','dest','doc','unidades','base','itens']} for p in payloads]}
    if results is not None: data['results']=results
    (OUTDIR/'manifest.json').write_text(json.dumps(data,indent=2,ensure_ascii=False))
    with (OUTDIR/'summary.csv').open('w',newline='') as f:
        w=csv.writer(f); w.writerow(['estab','dest','doc','unidades','base','itens','id','numero','situacao','chave','valorNota','pdf_path'])
        rows=results if results is not None else payloads
        for r in rows:
            latest=((r.get('latest') or {}).get('body') or {}).get('data') if isinstance((r.get('latest') or {}).get('body'),dict) else {}
            w.writerow([r.get('estab'),r.get('dest'),r.get('doc'),r.get('unidades'),r.get('base'),r.get('itens'),r.get('id'),r.get('numero') or latest.get('numero'),r.get('situacao') or latest.get('situacao'),r.get('chaveAcesso') or latest.get('chaveAcesso'),r.get('valorNota') or latest.get('valorNota'),r.get('pdf_path')])

def create_notes(payloads):
    results=[]
    for p in payloads:
        r=api(p['estab'],'POST','/nfe',p['payload'])
        rec={k:p[k] for k in ['estab','dest','doc','unidades','base','itens']}
        rec.update({'create_status':r['status'],'create_ok':r['ok'],'create_body':r['body']})
        if r['ok']:
            rec['id']=(r['body'].get('data') or {}).get('id')
            if rec.get('id'): rec['latest']=api(p['estab'],'GET',f'/nfe/{rec["id"]}')
        results.append(rec); time.sleep(1)
    return results

def send_notes(results):
    for rec in results:
        if not rec.get('id'): continue
        s=api(rec['estab'],'POST',f'/nfe/{rec["id"]}/enviar')
        rec['send_status']=s['status']; rec['send_ok']=s['ok']; rec['send_body']=s['body']
        time.sleep(3)
        rec['latest']=api(rec['estab'],'GET',f'/nfe/{rec["id"]}')
    return results

def download_pdfs(results):
    for rec in results:
        latest=((rec.get('latest') or {}).get('body') or {}).get('data') if isinstance((rec.get('latest') or {}).get('body'),dict) else {}
        if not latest: continue
        rec['numero']=latest.get('numero'); rec['situacao']=latest.get('situacao'); rec['chaveAcesso']=latest.get('chaveAcesso'); rec['valorNota']=latest.get('valorNota')
        link=latest.get('linkPDF') or latest.get('linkDanfe')
        if link:
            try:
                pdf=urllib.request.urlopen(link,timeout=30).read()
                path=OUTDIR/f"{rec['estab'].upper()}-NF-{rec['numero']}-{rec['dest'].replace(' ','_')}.pdf"
                path.write_bytes(pdf)
                rec['pdf_path']=str(path)
            except Exception as ex:
                rec['pdf_error']=str(ex)[:200]
    return results

if __name__=='__main__':
    mode=sys.argv[1] if len(sys.argv)>1 else 'preview'
    payloads=build_payloads(); write_summary(payloads)
    print(json.dumps({'mode':mode,'count':len(payloads),'total_unidades':sum(p['unidades'] for p in payloads),'total_base':round(sum(p['base'] for p in payloads),2),'preview':[{k:p[k] for k in ['estab','dest','unidades','base','itens']} for p in payloads]},ensure_ascii=False,indent=2))
    if mode=='create':
        res=create_notes(payloads); write_summary(payloads,res); print(json.dumps(res,ensure_ascii=False,indent=2))
    elif mode=='send':
        res=create_notes(payloads)
        if not all(r.get('id') for r in res):
            write_summary(payloads,res); print(json.dumps(res,ensure_ascii=False,indent=2)); sys.exit(2)
        res=send_notes(res); res=download_pdfs(res); write_summary(payloads,res)
        print(json.dumps([{k:r.get(k) for k in ['estab','dest','id','numero','situacao','send_status','send_ok','chaveAcesso','valorNota','pdf_path','pdf_error']} for r in res],ensure_ascii=False,indent=2))
