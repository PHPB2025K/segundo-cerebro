"""Preenche a planilha DFC POR CNPJ - [MÊS] 2026.xlsx (1 aba por dia).

Regras críticas (Knowledge File v10.1):
- Aba: nome `{DD:02d} de {Mês}` (ex: "21 de Maio")
- Saldo inicial: B73:I73 (vazias, editáveis APESAR da cor cinza)
- Lançamentos: linhas mapeadas em CAT_LINHA_CNPJ
- Transferências mantêm sinal; demais valores são positivos
- NUNCA tocar em J73 (fórmula) nem nas linhas de fórmula listadas
- Se mesma categoria aparece N vezes no mesmo dia/empresa → SOMAR antes de inserir
"""

import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from _constants import (
    D, MAPA_COLUNAS, LINHAS_FORMULA_CNPJ, MESES_PT,
)


def aba_do_dia(dia: int, mes_num: int) -> str:
    """Retorna nome da aba no formato {DD:02d} de {Mês} (Maiúscula primeira)."""
    return f"{dia:02d} de {MESES_PT[mes_num - 1]}"


def assert_celula_editavel(linha: int, coluna: int):
    """Garante que NÃO vamos escrever em célula de fórmula."""
    if linha in LINHAS_FORMULA_CNPJ:
        raise AssertionError(f"Linha {linha} é fórmula — não preencher")
    if linha == 73 and coluna == 10:  # J73
        raise AssertionError("J73 é fórmula (=SUM(B73:I73)) — não preencher")
    if coluna == 10:  # toda coluna J é consolidado
        raise AssertionError(f"Coluna J (consolidado) é fórmula — não preencher")


def preencher_dia(
    planilha_path: Path,
    dia: int,
    mes_num: int,
    saldos_iniciais: dict[str, Decimal],
    lancamentos_agregados: dict[tuple[str, int], Decimal],
):
    """Preenche uma aba (1 dia) da DFC POR CNPJ.

    Args:
        planilha_path: caminho da planilha DFC POR CNPJ - MES 2026.xlsx
        dia: dia do mês
        mes_num: número do mês (1-12)
        saldos_iniciais: {empresa: saldo_inicial} — empresas com movimento no dia
        lancamentos_agregados: {(empresa, linha): valor_acumulado_com_sinal}
            — para transferências o valor mantém sinal; demais entradas/saídas
              vêm como positivo (entrada) ou positivo absoluto (saída).
    """
    wb = openpyxl.load_workbook(planilha_path)
    nome_aba = aba_do_dia(dia, mes_num)
    if nome_aba not in wb.sheetnames:
        raise ValueError(f"Aba '{nome_aba}' não existe na planilha {planilha_path.name}")
    ws = wb[nome_aba]

    # Saldos iniciais em B73:I73
    for empresa, saldo in saldos_iniciais.items():
        if empresa not in MAPA_COLUNAS:
            raise ValueError(f"Empresa desconhecida: {empresa}")
        col = MAPA_COLUNAS[empresa]
        assert_celula_editavel(73, col)
        ws.cell(row=73, column=col, value=float(saldo))

    # Lançamentos agregados
    for (empresa, linha), valor in lancamentos_agregados.items():
        if empresa not in MAPA_COLUNAS:
            raise ValueError(f"Empresa desconhecida: {empresa}")
        col = MAPA_COLUNAS[empresa]
        assert_celula_editavel(linha, col)
        ws.cell(row=linha, column=col, value=float(valor))

    wb.save(planilha_path)


def agregar_lancamentos(extratos_processados: list) -> dict:
    """Agrega lançamentos por (dia, empresa, linha) somando valores.

    Transferências (linha 69) mantém sinal; entradas/saídas vão como valor
    absoluto positivo (a planilha não usa sinal nessas linhas — o sinal vem
    da categoria estrutural).

    Retorna: { dia_iso: { (empresa, linha): valor } }
    """
    agregados: dict[str, dict[tuple[str, int], Decimal]] = defaultdict(lambda: defaultdict(lambda: D("0")))
    for extrato in extratos_processados:
        if "erro" in extrato:
            continue
        empresa = extrato["empresa"]
        for lanc in extrato.get("lancamentos", []):
            cl = lanc.get("classificacao") or {}
            linha = cl.get("linha_cnpj")
            if linha is None:
                continue
            valor = D(lanc["valor"])
            if linha == 69:
                # transferência: sinal direto
                valor_signed = valor if lanc["sinal"] == "+" else -valor
                agregados[lanc["data"]][(empresa, 69)] += valor_signed
            else:
                # outras linhas: sempre positivo (sinal é estrutural)
                agregados[lanc["data"]][(empresa, linha)] += valor
    return agregados


if __name__ == "__main__":
    import json
    if len(sys.argv) != 4:
        print("Uso: python fill_dfc_por_cnpj.py <planilha.xlsx> <extratos_classificados.json> <mes_num>")
        sys.exit(1)

    planilha = Path(sys.argv[1])
    data = json.loads(Path(sys.argv[2]).read_text())
    mes_num = int(sys.argv[3])

    agregados = agregar_lancamentos(data)
    print(f"📋 Dias com movimento: {len(agregados)}")
    # TODO: saldos iniciais por dia → derivados via encadeamento
    # Este script preenche os lançamentos; preencher saldo inicial requer
    # cadeia partindo do SALDO ANTERIOR do extrato.
    print("⚠️ Saldos iniciais ainda não preenchidos automaticamente — usar")
    print("   preencher_dia() em script orquestrador que mantém o encadeamento.")
