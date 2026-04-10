#!/usr/bin/env python3
"""
Extrato Financeiro — Mercado Livre / Mercado Pago
Formato padrão GB Importadora / Budamix

Gera planilha Excel com todas as movimentações financeiras do período,
incluindo vendas, taxas, comissões, fretes, devoluções e disputas.

Uso:
    python3 ml-extrato.py --inicio 2026-03-08 --fim 2026-03-14
    python3 ml-extrato.py --inicio 2026-03-08 --fim 2026-03-14 --output /tmp/extrato.xlsx
"""

import json
import os
import sys
import argparse
import urllib.request
from datetime import datetime
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: apt-get install -y python3-openpyxl")
    sys.exit(1)

# ─── Configuração ────────────────────────────────────────────────────────────

TOKEN_FILE = "/root/.openclaw/.ml-tokens-finance.json"
USER_ID = 532562281
SELLER_NAME = "GB Importadora / Budamix"
BRAND_COLOR = "174C4C"
DEFAULT_OUTPUT_DIR = os.path.expanduser("~/.openclaw/workspace/reports")

# ─── Cores do resumo ─────────────────────────────────────────────────────────

GREEN = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BLUE = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
YELLOW = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
HEADER_FILL = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# ─── Classificação de transações ─────────────────────────────────────────────

LABELS = {
    'SETTLEMENT': 'Venda Concluída',
    'SETTLEMENT_SHIPPING': 'Repasse de Frete',
    'REFUND': 'Devolução / Reembolso',
    'REFUND_SHIPPING': 'Estorno de Frete',
    'DISPUTE': 'Reclamação / Disputa',
}


def classify_transaction(payment):
    """Classifica uma transação com base no status e descrição."""
    status = payment.get('status', '')
    status_detail = payment.get('status_detail', '')
    desc = (payment.get('description', '') or '').lower()
    amount = payment.get('transaction_amount', 0)
    is_shipping = 'shipment' in desc
    is_refund = status == 'refunded' or amount < 0

    if is_refund and is_shipping:
        return 'REFUND_SHIPPING'
    elif is_refund:
        return 'REFUND'
    elif is_shipping:
        return 'SETTLEMENT_SHIPPING'
    elif status in ('rejected', 'cancelled') or status_detail in ('bpp_refunded', 'mediation', 'chargeback'):
        return 'DISPUTE'
    return 'SETTLEMENT'


def get_liquidation_status(payment):
    """Determina se o pagamento já foi liquidado."""
    date_approved = payment.get('date_approved', '') or ''
    status = payment.get('status', '')
    if date_approved.strip():
        return 'Recebido'
    elif status in ('rejected', 'cancelled', 'refunded'):
        return 'Cancelado'
    return 'Pendente'


# ─── API ─────────────────────────────────────────────────────────────────────

def load_token():
    """Carrega o access token do arquivo."""
    if not os.path.exists(TOKEN_FILE):
        print(f"ERRO: Token não encontrado em {TOKEN_FILE}")
        print("Execute: ml-refresh-token.sh finance")
        sys.exit(1)
    with open(TOKEN_FILE) as f:
        return json.load(f)['access_token']


def fetch_payments(token, begin_date, end_date):
    """Busca todos os pagamentos no período via API do Mercado Pago."""
    all_payments = []
    offset = 0
    limit = 50

    while True:
        url = (
            f"https://api.mercadopago.com/v1/payments/search"
            f"?begin_date={begin_date}T00:00:00Z"
            f"&end_date={end_date}T23:59:59Z"
            f"&limit={limit}&offset={offset}"
            f"&sort=date_created&criteria=desc"
        )
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        
        try:
            resp = urllib.request.urlopen(req)
            data = json.loads(resp.read())
        except Exception as e:
            print(f"ERRO na API (offset={offset}): {e}")
            if all_payments:
                print(f"Continuando com {len(all_payments)} pagamentos já obtidos...")
                break
            sys.exit(1)

        results = data.get('results', [])
        all_payments.extend(results)
        total = data.get('paging', {}).get('total', 0)

        if offset + limit >= total or not results:
            break
        offset += limit

    return all_payments


# ─── Processamento ───────────────────────────────────────────────────────────

def process_payment(p):
    """Converte um pagamento da API em uma linha do extrato."""
    pid = p.get('id', '')
    order_id = p.get('order', {}).get('id', '') if p.get('order') else ''
    amount = p.get('transaction_amount', 0)
    net = p.get('transaction_details', {}).get('net_received_amount', 0)
    fee = round(amount - net, 2)
    date_created = p.get('date_created', '')[:19].replace('T', ' ')
    date_approved = (p.get('date_approved', '') or '')[:19].replace('T', ' ')
    
    # Shipping ID from charges
    shipping_id = ''
    charges = p.get('charges_details', [])
    for c in charges:
        meta = c.get('metadata', {})
        if meta.get('shipment_id'):
            shipping_id = str(meta['shipment_id'])

    metadata_str = json.dumps(charges[:1]) if charges else '[{}]'
    if len(metadata_str) > 100:
        metadata_str = metadata_str[:100] + '...'

    tx_type = classify_transaction(p)

    return {
        'ref_ext': order_id or pid,
        'id_origem': pid,
        'id_usuario': USER_ID,
        'tipo_pagamento': p.get('payment_type_id', ''),
        'metodo': p.get('payment_method_id', ''),
        'site': 'MLB',
        'tipo_transacao': tx_type,
        'classificacao': LABELS.get(tx_type, tx_type),
        'valor': round(amount, 2),
        'moeda': p.get('currency_id', 'BRL'),
        'data_transacao': date_created,
        'taxas': round(-abs(fee), 2) if fee > 0 else round(fee, 2),
        'valor_liquido': round(net, 2),
        'moeda_liq': p.get('currency_id', 'BRL'),
        'data_liquidacao': date_approved,
        'status_liquidacao': get_liquidation_status(p),
        'valor_real': round(net, 2),
        'cupom': p.get('coupon_amount', 0),
        'metadata': metadata_str,
        'id_pedido': order_id,
        'id_envio': shipping_id,
        'modo_envio': 'me2',
        'id_pack': ''
    }


# ─── Excel ───────────────────────────────────────────────────────────────────

HEADERS = [
    "Ref. Externa", "ID Origem", "ID Usuário", "Tipo Pagamento", "Método", "Site",
    "Tipo (técnico)", "Classificação", "Valor (R$)", "Moeda", "Data Transação", "Taxas (R$)",
    "Valor Líquido (R$)", "Moeda Liq.", "Data Liquidação", "Status Liquidação",
    "Valor Real (R$)", "Cupom (R$)", "Metadata", "ID Pedido", "ID Envio", "Modo Envio", "ID Pack"
]

ROW_KEYS = [
    'ref_ext', 'id_origem', 'id_usuario', 'tipo_pagamento', 'metodo', 'site',
    'tipo_transacao', 'classificacao', 'valor', 'moeda', 'data_transacao', 'taxas',
    'valor_liquido', 'moeda_liq', 'data_liquidacao', 'status_liquidacao',
    'valor_real', 'cupom', 'metadata', 'id_pedido', 'id_envio', 'modo_envio', 'id_pack'
]

COL_WIDTHS = {
    'A':18, 'B':16, 'C':12, 'D':18, 'E':18, 'F':6, 'G':22, 'H':22,
    'I':12, 'J':6, 'K':22, 'L':12, 'M':14, 'N':8, 'O':22, 'P':16,
    'Q':12, 'R':8, 'S':30, 'T':18, 'U':14, 'V':10, 'W':14
}


def build_excel(rows, begin_date, end_date, output_path):
    """Gera o arquivo Excel no formato padrão."""
    wb = Workbook()
    ws = wb.active
    
    # Format dates for display
    d1 = datetime.strptime(begin_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    d2 = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    month_name = datetime.strptime(begin_date, '%Y-%m-%d').strftime('%B %Y').upper()
    
    ws.title = f"Extrato {begin_date[8:10]}-{end_date[8:10]} {datetime.strptime(begin_date, '%Y-%m-%d').strftime('%b')}"

    # ── Cabeçalho ──
    ws.cell(row=1, column=1,
            value=f"EXTRATO FINANCEIRO — {d1} a {d2} — {month_name} — MERCADO LIVRE"
            ).font = Font(bold=True, size=14, color=BRAND_COLOR)
    ws.cell(row=2, column=1,
            value=f"{SELLER_NAME}  •  User ID: {USER_ID}  •  Período: {d1} a {d2}"
            ).font = Font(size=11, color="666666")

    # ── Headers ──
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # ── Data ──
    for i, r in enumerate(rows, 5):
        for col, key in enumerate(ROW_KEYS, 1):
            cell = ws.cell(row=i, column=col, value=r[key])
            cell.border = THIN_BORDER

    last_row = 4 + len(rows)
    s = last_row + 2

    # ── Resumo ──
    ws.cell(row=s, column=1, value="RESUMO FINANCEIRO").font = Font(bold=True, size=13, color=BRAND_COLOR)
    ws.cell(row=s, column=3, value="% do Bruto").font = Font(bold=True, size=11, color=BRAND_COLOR)

    bruto_formula = f'SUMPRODUCT((G5:G{last_row}="SETTLEMENT")*(I5:I{last_row}))'
    bold_font = Font(bold=True, size=11)
    pct_font = Font(bold=True, size=11, color="555555")
    desc_font = Font(italic=True, color="888888")

    summary_lines = [
        ("💰 Total de Vendas", f'=SUMPRODUCT((G5:G{last_row}="SETTLEMENT")*(I5:I{last_row}))', GREEN, True, "Soma de todas as vendas concluídas no período"),
        ("📦 Frete Repassado", f'=SUMPRODUCT((G5:G{last_row}="SETTLEMENT_SHIPPING")*(I5:I{last_row}))', YELLOW, True, "Frete cobrado do comprador e repassado à transportadora"),
        ("🔄 Devoluções / Reembolsos", f'=SUMPRODUCT((G5:G{last_row}="REFUND")*(I5:I{last_row}))', RED, True, "Valor devolvido ao comprador por cancelamento ou devolução"),
        ("⚠️ Reclamações / Disputas", f'=SUMPRODUCT((G5:G{last_row}="DISPUTE")*(I5:I{last_row}))', RED, True, "Valor retido ou devolvido por reclamação do comprador"),
        ("🚚 Estorno de Frete", f'=SUMPRODUCT((G5:G{last_row}="REFUND_SHIPPING")*(I5:I{last_row}))', RED, True, "Frete estornado em vendas canceladas/devolvidas"),
        ("", "", None, False, ""),
        ("📊 Total de Taxas e Comissões ML", f'=SUM(L5:L{last_row})', RED, True, "Tudo que o ML cobrou: comissão, processamento, frete"),
        ("", "", None, False, ""),
        ("Qtd. Vendas Concluídas", f'=COUNTIF(G5:G{last_row},"SETTLEMENT")', None, False, ""),
        ("Qtd. Devoluções", f'=COUNTIF(G5:G{last_row},"REFUND")', None, False, ""),
        ("Qtd. Disputas", f'=COUNTIF(G5:G{last_row},"DISPUTE")', None, False, ""),
        ("Qtd. Fretes", f'=COUNTIF(G5:G{last_row},"SETTLEMENT_SHIPPING")', None, False, ""),
        ("", "", None, False, ""),
        ("Recebido (já caiu no Mercado Pago)", f'=SUMPRODUCT((P5:P{last_row}="Recebido")*(M5:M{last_row}))', GREEN, True, "Valor já disponível na conta do Mercado Pago"),
        ("Pendente (ainda não caiu)", f'=SUMPRODUCT((P5:P{last_row}="Pendente")*(M5:M{last_row}))', YELLOW, True, "Valor ainda em processamento"),
        ("", "", None, False, ""),
        ("✅ SALDO LÍQUIDO FINAL", f'=SUM(M5:M{last_row})', BLUE, True, "O que sobrou pra você depois de tudo"),
    ]

    for j, (label, formula, fill, has_pct, desc) in enumerate(summary_lines):
        r = s + 1 + j
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = bold_font
        if fill:
            c1.fill = fill

        c2 = ws.cell(row=r, column=2, value=formula if formula else "")
        c2.font = Font(bold=True)
        c2.number_format = '#,##0.00'
        if fill:
            c2.fill = fill

        if has_pct and label:
            c3 = ws.cell(row=r, column=3, value=f'=B{r}/{bruto_formula}')
            c3.font = pct_font
            c3.number_format = '0.0%'
            c3.alignment = Alignment(horizontal="center")
            if fill:
                c3.fill = fill

        if desc:
            ws.cell(row=r, column=5, value=desc).font = desc_font

    # ── Larguras ──
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Salvar ──
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extrato Financeiro ML/MP")
    parser.add_argument("--inicio", required=True, help="Data inicial (YYYY-MM-DD)")
    parser.add_argument("--fim", required=True, help="Data final (YYYY-MM-DD)")
    parser.add_argument("--output", help="Caminho de saída do Excel")
    args = parser.parse_args()

    if not args.output:
        d1 = args.inicio[5:].replace('-', '')
        d2 = args.fim[5:].replace('-', '')
        args.output = os.path.join(DEFAULT_OUTPUT_DIR, f"extrato-ml-{d1}-a-{d2}.xlsx")

    print(f"📊 Extrato ML — {args.inicio} a {args.fim}")
    print(f"   Buscando pagamentos...")

    token = load_token()
    payments = fetch_payments(token, args.inicio, args.fim)
    print(f"   {len(payments)} pagamentos encontrados")

    rows = [process_payment(p) for p in payments]

    types = Counter(r['tipo_transacao'] for r in rows)
    liq = Counter(r['status_liquidacao'] for r in rows)
    print(f"   Tipos: {dict(types)}")
    print(f"   Liquidação: {dict(liq)}")

    output = build_excel(rows, args.inicio, args.fim, args.output)
    print(f"   ✅ Salvo em: {output}")


if __name__ == "__main__":
    main()
