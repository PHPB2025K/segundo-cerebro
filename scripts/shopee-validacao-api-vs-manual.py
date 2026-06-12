#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""shopee-validacao-api-vs-manual.py — Valida a integração Shopee API (OpenClaw)
contra extrações manuais do Seller Center.

Criado na validação de 12/06/2026 (conta budamix-store 448649947, períodos
10/05–10/06/2026). Resultado daquela rodada: 100% de match em pedidos
(order_sn), extrato de renda (liberações, R$ 0,00 de diferença) e amostra
campo a campo (150/150). Diferenças conhecidas e esperadas (NÃO são erro):
  - Painel BI "Vendas": base de cálculo própria do painel ≠ total_amount da
    API (~3-4%). Contagem de pedidos pagos bate exato.
  - "Total global" do export de pedidos: zerado em Cancelados e
    Devolvidos/Reembolsados; API mantém valor original do pedido.
  - Shopee Ads: app da integração SEM permissão na API de Ads (403) —
    gasto de Ads só via export manual do Seller Center.

SOMENTE LEITURA na API (usa o access_token corrente, sem refresh — o serviço
shopee_oauth mantém os tokens frescos).

USO MENSAL (2 passos):

1) Na VPS — coletar dados da API (datas em BRT, formato YYYY-MM-DD):
   python3 shopee-validacao-api-vs-manual.py fetch \
     --shop-id 448649947 \
     --orders-from 2026-06-01 --orders-to 2026-06-30 \
     --escrow-list-from 2026-06-01 --escrow-list-to 2026-06-30 \
     --output /tmp/shopee-api-mes.json

2) Comparar com os exports manuais do Seller Center (xlsx):
   python3 shopee-validacao-api-vs-manual.py compare \
     --pedidos "PEDIDOS - SHOPEE 1 - 30 DIAS.xlsx" \
     --renda "RENDA - SHOPEE 1 - 30 DIAS.xlsx" \
     --api /tmp/shopee-api-mes.json \
     [--escrow-sample 15]   # refaz amostra campo a campo (chama a API)

   Exports esperados: "Meus Pedidos" (aba orders) e "Minha Renda" (abas
   Summary + Renda), ambos do Seller Center. Períodos: pedidos por DATA DE
   CRIAÇÃO, renda por DATA DE LIBERAÇÃO — exporte com as mesmas janelas
   usadas no fetch.

Contas (integrations/shopee): 448649947 budamix-store | 860803675
budamix-oficial-2 | 442066454 budamix-shop-3.
"""

import argparse
import hashlib
import hmac
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

HOST = "https://partner.shopeemobile.com"
PARTNER_ID = 2031533
PARTNER_KEY_PATH = "/root/.openclaw/workspace/integrations/shopee/.env"
TOKENS_DIR = "/root/.openclaw/workspace/integrations/shopee"
SHOP_ID_TO_TOKEN_FILE = {
    448649947: ".shopee-tokens-budamix-store.json",
    860803675: ".shopee-tokens-budamix-store2.json",
    442066454: ".shopee-tokens-budamix-shop.json",
}
BRT = timezone(timedelta(hours=-3))
HTTP_TIMEOUT = 30
TOL_TOTAL = 0.05   # tolerância em totais
TOL_ORDER = 0.01   # tolerância por pedido


# ─── API helpers ──────────────────────────────────────────────────────────────

def load_partner_key():
    for line in open(PARTNER_KEY_PATH).read().splitlines():
        if line.startswith("SHOPEE_PARTNER_KEY="):
            return line.split("=", 1)[1].strip()
    sys.exit("SHOPEE_PARTNER_KEY ausente")


def sign(key, path, ts, access_token="", shop_id=0):
    base = f"{PARTNER_ID}{path}{ts}"
    if access_token:
        base += f"{access_token}{shop_id}"
    return hmac.new(key.encode(), base.encode(), hashlib.sha256).hexdigest()


def api_get(key, path, params, shop_id, access_token):
    ts = int(time.time())
    full = {
        "partner_id": str(PARTNER_ID),
        "timestamp": str(ts),
        "sign": sign(key, path, ts, access_token, shop_id),
        "shop_id": str(shop_id),
        "access_token": access_token,
        **{k: str(v) for k, v in params.items()},
    }
    url = f"{HOST}{path}?{urllib.parse.urlencode(full)}"
    for attempt in range(5):
        try:
            resp = urllib.request.urlopen(urllib.request.Request(url), timeout=HTTP_TIMEOUT)
            data = json.loads(resp.read())
            if data.get("error") and "too many" in str(data.get("message", "")).lower():
                time.sleep(2 ** attempt)
                continue
            return data
        except Exception as e:
            if attempt == 4:
                return {"error": "exception", "message": str(e)}
            time.sleep(1 + attempt)
    return {"error": "max_retries"}


def load_access_token(shop_id):
    with open(f"{TOKENS_DIR}/{SHOP_ID_TO_TOKEN_FILE[shop_id]}") as f:
        return json.load(f)["access_token"]


def brt_range_to_unix(date_from, date_to):
    t0 = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=BRT)
    t1 = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=BRT) + timedelta(days=1) - timedelta(seconds=1)
    return int(t0.timestamp()), int(t1.timestamp())


def brt_date(unix):
    return datetime.fromtimestamp(unix, BRT).strftime("%Y-%m-%d")


# ─── fetch ────────────────────────────────────────────────────────────────────

def fetch_order_list(key, shop_id, token, unix_from, unix_to):
    out, win_from = [], unix_from
    while win_from <= unix_to:
        win_to = min(win_from + 15 * 86400 - 1, unix_to)
        cursor = ""
        while True:
            params = {
                "time_range_field": "create_time",
                "time_from": win_from, "time_to": win_to,
                "page_size": 100, "response_optional_fields": "order_status",
            }
            if cursor:
                params["cursor"] = cursor
            data = api_get(key, "/api/v2/order/get_order_list", params, shop_id, token)
            if data.get("error"):
                sys.exit(f"get_order_list: {data}")
            resp = data.get("response", {})
            out.extend({"order_sn": o["order_sn"], "order_status": o.get("order_status")}
                       for o in resp.get("order_list", []))
            if resp.get("more") and resp.get("next_cursor"):
                cursor = resp["next_cursor"]
            else:
                break
        win_from = win_to + 1
    return out


def fetch_order_details(key, shop_id, token, order_sns):
    fields = ("order_status,total_amount,create_time,pay_time,buyer_username,"
              "item_list,payment_method,order_sn,actual_shipping_fee,estimated_shipping_fee")
    out = {}
    for i in range(0, len(order_sns), 50):
        batch = order_sns[i:i + 50]
        data = api_get(key, "/api/v2/order/get_order_detail",
                       {"order_sn_list": ",".join(batch), "response_optional_fields": fields},
                       shop_id, token)
        if data.get("error"):
            sys.exit(f"get_order_detail (lote {i}): {data}")
        for o in data.get("response", {}).get("order_list", []):
            out[o["order_sn"]] = {
                "order_status": o.get("order_status"),
                "total_amount": o.get("total_amount"),
                "create_time": o.get("create_time"),
                "pay_time": o.get("pay_time"),
                "items": [{"item_sku": it.get("item_sku"), "model_sku": it.get("model_sku"),
                           "qty": it.get("model_quantity_purchased"),
                           "price": it.get("model_discounted_price")}
                          for it in (o.get("item_list") or [])],
            }
        time.sleep(0.15)
    return out


def fetch_escrow_list(key, shop_id, token, unix_from, unix_to):
    out, page_no = [], 1
    while True:
        data = api_get(key, "/api/v2/payment/get_escrow_list",
                       {"release_time_from": unix_from, "release_time_to": unix_to,
                        "page_size": 100, "page_no": page_no},
                       shop_id, token)
        if data.get("error"):
            sys.exit(f"get_escrow_list (pág {page_no}): {data}")
        resp = data.get("response", {})
        rows = resp.get("escrow_list", [])
        out.extend(rows)
        if (not resp.get("more") and len(rows) < 100) or not rows:
            break
        page_no += 1
        time.sleep(0.12)
    return out


def fetch_escrow_detail(key, shop_id, token, order_sn):
    data = api_get(key, "/api/v2/payment/get_escrow_detail", {"order_sn": order_sn}, shop_id, token)
    return data.get("response", {}) if not data.get("error") else {"error": data}


def cmd_fetch(args):
    key = load_partner_key()
    token = load_access_token(args.shop_id)
    result = {"shop_id": args.shop_id, "fetched_at": datetime.now(BRT).isoformat(),
              "orders_window": [args.orders_from, args.orders_to],
              "escrow_window": [args.escrow_list_from, args.escrow_list_to]}

    if args.orders_from:
        uf, ut = brt_range_to_unix(args.orders_from, args.orders_to)
        orders = fetch_order_list(key, args.shop_id, token, uf, ut)
        print(f"[orders] {len(orders)} pedidos", file=sys.stderr)
        result["orders"] = orders
        result["order_details"] = fetch_order_details(
            key, args.shop_id, token, [o["order_sn"] for o in orders])
        print(f"[details] {len(result['order_details'])} detalhados", file=sys.stderr)

    if args.escrow_list_from:
        uf, ut = brt_range_to_unix(args.escrow_list_from, args.escrow_list_to)
        result["escrow_list"] = fetch_escrow_list(key, args.shop_id, token, uf, ut)
        print(f"[escrow_list] {len(result['escrow_list'])} liberações", file=sys.stderr)

    with open(args.output, "w") as f:
        json.dump(result, f, ensure_ascii=False)
    print(f"OK -> {args.output}", file=sys.stderr)


# ─── compare ──────────────────────────────────────────────────────────────────

def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_pedidos_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["orders"]
    orders = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        sn = r[0]
        if not sn:
            continue
        if sn not in orders:
            orders[sn] = {"status": r[1], "create": str(r[8] or ""),
                          "total_global": r[51]}
    wb.close()
    return orders


def read_renda_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Renda"]
    rows = {}
    for r in ws.iter_rows(min_row=4, max_col=35, values_only=True):
        if r[1] != "Order":
            continue
        rows[r[2]] = {
            "lancado": r[11], "preco_produto": r[12], "reembolso": r[13],
            "data_liberacao": str(r[7] or ""),
            "frete_comprador": r[16], "frete_parceiro": r[17],
            "desc_frete_shopee": r[18], "comissao_liq": r[27],
            "servico_liq": r[28], "afiliados": r[30], "pago_comprador": r[34],
        }
    wb.close()
    return rows


ESCROW_MAP = [
    ("Valor liberado",        "lancado",          "escrow_amount"),
    ("Preço do produto",      "preco_produto",    "cost_of_goods_sold"),
    ("Comissão líquida",      "comissao_liq",     "net_commission_fee"),
    ("Taxa serviço líquida",  "servico_liq",      "net_service_fee"),
    ("Frete pago comprador",  "frete_comprador",  "buyer_paid_shipping_fee"),
    ("Frete parceiro logíst", "frete_parceiro",   "actual_shipping_fee"),
    ("Desc. frete Shopee",    "desc_frete_shopee","shopee_shipping_rebate"),
    ("Comissão afiliados",    "afiliados",        "order_ams_commission_fee"),
    ("Pago pelo comprador",   "pago_comprador",   "buyer_total_amount"),
    ("Reembolso",             "reembolso",        "seller_return_refund"),
]


def cmd_compare(args):
    man_ped = read_pedidos_xlsx(args.pedidos)
    man_renda = read_renda_xlsx(args.renda)
    api = json.load(open(args.api))
    api_orders = {o["order_sn"] for o in api.get("orders", [])}
    api_det = api.get("order_details", {})
    api_esc = api.get("escrow_list", [])
    api_esc_by_sn = {}
    for e in api_esc:
        api_esc_by_sn[e["order_sn"]] = api_esc_by_sn.get(e["order_sn"], 0) + e.get("payout_amount", 0)

    print("=" * 66)
    print("NÍVEL 1 — TOTAIS")
    print("=" * 66)
    print(f"Pedidos:  manual {len(man_ped)} × API {len(api_orders)} "
          f"{'✅' if len(man_ped) == len(api_orders) else '❌'}")
    man_total = sum(_f(v["lancado"]) or 0 for v in man_renda.values())
    api_total = sum(api_esc_by_sn.values())
    ok = abs(man_total - api_total) <= TOL_TOTAL
    print(f"Renda:    manual {len(man_renda)} ped / R$ {man_total:,.2f} × "
          f"API {len(api_esc_by_sn)} ped / R$ {api_total:,.2f} {'✅' if ok else '❌'}")

    print()
    print("=" * 66)
    print("NÍVEL 2 — POR PEDIDO (order_sn)")
    print("=" * 66)
    for label, mset, aset, mdates in [
        ("Pedidos", set(man_ped), api_orders,
         {sn: v["create"][:10] for sn, v in man_ped.items()}),
        ("Renda", set(man_renda), set(api_esc_by_sn),
         {sn: v["data_liberacao"] for sn, v in man_renda.items()}),
    ]:
        so_m, so_a = sorted(mset - aset), sorted(aset - mset)
        print(f"{label}: comum {len(mset & aset)} | só manual {len(so_m)} | só API {len(so_a)}")
        for sn in so_m[:15]:
            print(f"   só MANUAL: {sn} | {mdates.get(sn, '?')}")
        for sn in so_a[:15]:
            d = api_det.get(sn, {})
            dt = brt_date(d["create_time"]) if d.get("create_time") else "?"
            print(f"   só API:    {sn} | {dt}")

    diffs = [(sn, _f(v["lancado"]), api_esc_by_sn[sn])
             for sn, v in man_renda.items()
             if sn in api_esc_by_sn and _f(v["lancado"]) is not None
             and abs(_f(v["lancado"]) - api_esc_by_sn[sn]) > TOL_ORDER]
    print(f"Renda por pedido: {len(set(man_renda) & set(api_esc_by_sn))} comparados | "
          f"{len(diffs)} divergentes (> R$ {TOL_ORDER:.2f}) {'✅' if not diffs else '❌'}")
    for sn, m, a in diffs[:20]:
        print(f"   {sn}: manual {m:.2f} × API {a:.2f}")

    # Total global × total_amount (ignora Cancelado/Devolvido — semântica do export)
    sem, real = 0, []
    for sn in set(man_ped) & set(api_det):
        mv, av = _f(man_ped[sn]["total_global"]), api_det[sn].get("total_amount")
        if mv is None or av is None or abs(mv - av) <= TOL_ORDER:
            continue
        if man_ped[sn]["status"] == "Cancelado" or _f(man_renda.get(sn, {}).get("reembolso")):
            sem += 1   # esperado: export zera valor em cancelado/devolvido
        else:
            real.append((sn, mv, av))
    print(f"Total global × total_amount: {sem} difs semânticas (cancelado/devolvido, esperadas) | "
          f"{len(real)} divergências reais {'✅' if not real else '❌'}")
    for sn, m, a in real[:15]:
        print(f"   ❌ {sn}: manual {m:.2f} × API {a:.2f}")

    if args.escrow_sample:
        print()
        print("=" * 66)
        print(f"NÍVEL 3 — AMOSTRA CAMPO A CAMPO ({args.escrow_sample} pedidos)")
        print("=" * 66)
        key = load_partner_key()
        token = load_access_token(api["shop_id"])
        pool = sorted(set(man_renda) & set(api_esc_by_sn))
        step = max(1, len(pool) // args.escrow_sample)
        sample = pool[::step][:args.escrow_sample]
        n_cmp, n_div = 0, 0
        for sn in sample:
            esc = fetch_escrow_detail(key, api["shop_id"], token, sn)
            oi = esc.get("order_income", {})
            mr = man_renda[sn]
            bad = []
            for label, mk, ak in ESCROW_MAP:
                mv, av = _f(mr.get(mk)) or 0.0, _f(oi.get(ak)) or 0.0
                n_cmp += 1
                if abs(abs(mv) - abs(av)) > TOL_ORDER:   # manual usa sinal negativo p/ taxas
                    n_div += 1
                    bad.append(f"{label}: manual {mv:.2f} × API {av:.2f}")
            print(f"  {'✅' if not bad else '❌'} {sn}" + ("" if not bad else " — " + "; ".join(bad)))
            time.sleep(0.25)
        print(f"\n{n_cmp} comparações, {n_div} divergências reais")

    print("\nLembretes: BI do Seller Center tem base própria de 'Vendas' (não usar "
          "como base financeira); Ads não é validável via API (app sem permissão).")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    pf = sub.add_parser("fetch", help="coletar dados da API Shopee")
    pf.add_argument("--shop-id", type=int, required=True)
    pf.add_argument("--orders-from")
    pf.add_argument("--orders-to")
    pf.add_argument("--escrow-list-from")
    pf.add_argument("--escrow-list-to")
    pf.add_argument("--output", required=True)
    pf.set_defaults(func=cmd_fetch)

    pc = sub.add_parser("compare", help="comparar exports manuais × JSON da API")
    pc.add_argument("--pedidos", required=True, help="xlsx Meus Pedidos (aba orders)")
    pc.add_argument("--renda", required=True, help="xlsx Minha Renda (aba Renda)")
    pc.add_argument("--api", required=True, help="JSON gerado pelo fetch")
    pc.add_argument("--escrow-sample", type=int, default=0,
                    help="N pedidos p/ amostra campo a campo via get_escrow_detail")
    pc.set_defaults(func=cmd_compare)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
